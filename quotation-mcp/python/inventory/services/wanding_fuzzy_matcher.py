"""
万鼎价格库匹配 - DataBase-style 模糊逻辑

仅此一种查询逻辑：token + 同义词扩展 + 规格等价 + score 排序。
借鉴 DataBase- 项目 search_with_keywords(strict=False, return_score=True)。
"""

from __future__ import annotations

import logging
import re
import threading
from pathlib import Path
from typing import Any, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# 价格列（0-based）：与价格库表头顺序一致；多档位共用同一列时取同一索引
PRICE_COLS = {
    "FACTORY_INC_TAX": 4,
    "FACTORY_EXC_TAX": 5,
    "PURCHASE_EXC_TAX": 6,
    # 下列索引基于「万鼎价格库_管材与国标管件_标准格式.xlsx」表头（含 Product_Type 列后偏移 +1）：
    # 列5-7: 出厂价；列8/9: A 档 利润率/报单价格；列10/11: B 档；列12/13: C 档；列14/15: D 档；列16/17: D 低利润率；列18/19: E 档。
    "A_MARGIN": 9,
    "A_QUOTE": 9,
    "B_MARGIN": 11,
    "B_QUOTE": 11,
    "C_MARGIN": 13,
    "C_QUOTE": 13,
    "D_MARGIN": 15,
    "D_QUOTE": 15,
    "D_LOW": 17,
    "E_MARGIN": 19,
    "E_QUOTE": 19,
    # 兼容旧代码
    "A": 9,
    "A_TURN": 9,
    "A_ANNUAL": 9,
    "B": 11,
    "B_TURN": 11,
    "B_ANNUAL": 11,
    "B_QUOTE": 11,
    "C": 13,
    "C_TURN": 13,
    "C_QUOTE": 13,
    "D": 15,
    "D_NOADJ": 15,
    "D_WHOLESALE": 17,
    "E": 19,
}

# 对应每个档位价格列的「利润率」列索引（0-based）。
PROFIT_COLS: dict[str, int] = {
    "A_MARGIN": 8,
    "A_QUOTE": 8,
    "B_MARGIN": 10,
    "B_QUOTE": 10,
    "C_MARGIN": 12,
    "C_QUOTE": 12,
    "D_MARGIN": 14,
    "D_QUOTE": 14,
    "D_LOW": 16,
    "E_MARGIN": 18,
    "E_QUOTE": 18,
}


def _normalize_price_level(customer_level: str) -> str:
    """将用户/Agent 传入的档位或价格类型统一为 PRICE_COLS 的 key。"""
    s = (customer_level or "B_QUOTE").strip()
    if not s:
        return "B_QUOTE"
    # 中文：出厂价含税/不含税、采购不含税
    if "出厂价" in s and "含税" in s:
        return "FACTORY_INC_TAX"
    if "出厂价" in s and "不含税" in s:
        return "FACTORY_EXC_TAX"
    if "采购" in s and "不含税" in s:
        return "PURCHASE_EXC_TAX"
    # 英文/代码
    u = s.upper().replace(" ", "_")
    if u in ("D_LOW", "D LOW", "DLOW"):
        return "D_LOW"
    for key in ("FACTORY_INC_TAX", "FACTORY_EXC_TAX", "PURCHASE_EXC_TAX"):
        if key in u or u == key:
            return key
    return u if u in PRICE_COLS else "B_QUOTE"


# 档位代码 → 全名（与价格库表头一致，Chat/Work 统一显示）
PRICE_LEVEL_DISPLAY_NAMES: dict[str, str] = {
    "FACTORY_INC_TAX": "出厂价_含税",
    "FACTORY_EXC_TAX": "出厂价_不含税",
    "PURCHASE_EXC_TAX": "采购不含税",
    "A_MARGIN": "（二级代理）A级别 利润率",
    "A_QUOTE": "（二级代理）A级别 报单价格",
    "B_MARGIN": "（一级代理）B级别 利润率",
    "B_QUOTE": "（一级代理）B级别 报单价格",
    "C_MARGIN": "（聚万大客户）C级别 利润率",
    "C_QUOTE": "（聚万大客户）C级别报单价格",
    "D_MARGIN": "（青山大客户）D级别 利润率",
    "D_QUOTE": "（青山大客户）D级别 报单价格",
    "D_LOW": "（青山大客户）D级别 降低利润率",
    "E_MARGIN": "（大唐大客户）E级别（包运费） 利润率",
    "E_QUOTE": "（大唐大客户）E级别（包运费） 报单价格",
    # 兼容旧代码
    "A": "（二级代理）A级别 利润率",
    "A_TURN": "（二级代理）A级别 利润率",
    "A_ANNUAL": "（二级代理）A级别 报单价格",
    "B": "（一级代理）B级别 利润率",
    "B_TURN": "（一级代理）B级别 利润率",
    "B_ANNUAL": "（一级代理）B级别 报单价格",
    "C": "（聚万大客户）C级别 利润率",
    "C_TURN": "（聚万大客户）C级别 利润率",
    "D": "（青山大客户）D级别 利润率",
    "D_NOADJ": "（青山大客户）D级别 报单价格",
    "D_WHOLESALE": "（青山大客户）D级别 降低利润率",
    "E": "（大唐大客户）E级别（包运费） 利润率",
}


def get_price_level_display_name(customer_level: str) -> str:
    """将档位代码转为全名，供界面与接口返回使用。"""
    key = _normalize_price_level(customer_level or "B_QUOTE")
    return PRICE_LEVEL_DISPLAY_NAMES.get(key, key)


SYNONYM_GROUPS = [
    {"直接", "直接头", "直通", "直通接头"},
    {"变径", "异径"},
    {"大小头", "异径直通", "异径套", "变径直接", "异径直接"},
    {"内丝", "内螺纹"}, {"外丝", "外螺纹"},
    {"锁母", "锁扣", "管接头"},
    {"止回阀", "截止阀"},
    {"穿线管", "电线管"},
    {"半弯", "弯头"},
    {"承插", "承插式"},
    {"堵头", "管帽"},
]

# 模块级预计算，避免每次调用时重建（SYNONYM_GROUPS 不变时永远有效）
_SYNONYM_TO_GROUP: dict[str, frozenset] = {
    syn: frozenset(group) for group in SYNONYM_GROUPS for syn in group
}
_SORTED_SYNONYMS: list[str] = sorted(_SYNONYM_TO_GROUP.keys(), key=len, reverse=True)

# 单字 token（如「三」「通」）在打分中的权重，相对于多字 token 的 1.0
_SINGLE_CHAR_WEIGHT = 0.5

# 询价关键词中的英文/印尼语/口语 → 中文品名，用于筛选时命中库内品名（与 wanding_business_knowledge.md 保持一致，便于 LLM 选型有思路）
QUERY_TERM_TO_CHINESE = [
    ("4 cabang", "管四通圆接线盒"),
    ("conduit", "电线管"), ("counduit", "电线管"), ("pipa", "管"),
    ("socket", "管直通"), ("套筒", "管直通"),
    ("klem", "管夹"),
    ("cabang", "四通"), ("tdust", "四通"),
    ("热熔器", "焊接机"), ("热熔机", "焊接机"), ("熔接器", "焊接机"),
    ("四通接线盒", "管四通圆接线盒"),
    ("马鞍卡", "管夹"),
    # AW 在印尼管材场景下常指 AW 给水系列
    ("aw", "给水 aw给水系列"),
]


def _normalize_keyword_terms(keywords: str) -> str:
    """将询价中的英文/印尼语替换为中文品名词，便于筛选命中库内品名。"""
    s = (keywords or "").strip()
    for eng, ch in QUERY_TERM_TO_CHINESE:
        s = re.sub(r"\b" + re.escape(eng) + r"\b", ch, s, flags=re.I)
    return s.strip()


# 询价意图词（价格/档位）不应参与产品字段匹配，否则会把正确候选过滤掉
_QUERY_INTENT_STOPWORDS = {
    "报价", "报单", "价格", "价", "报价价格",
    "一级", "二级", "三级", "代理", "代理价", "一级代理", "二级代理",
    "a级", "b级", "c级", "d级", "e级",
}


def _strip_query_intent_terms(keywords: str) -> str:
    """移除询价中的非品名意图词，保留材质/规格/品类 token 用于字段匹配。"""
    s = _normalize(keywords or "")
    if not s:
        return ""
    # 先按长词优先剔除，避免残留碎片（如先去「一级代理」再去「一级」）
    for term in sorted(_QUERY_INTENT_STOPWORDS, key=len, reverse=True):
        s = re.sub(re.escape(term), " ", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# 业务知识中【字段匹配同义与规格】规则缓存，供字段匹配阶段使用（与 LLM 选型共用同一 knowledge 文件）
_FIELD_MATCHING_RULES_CACHE: dict = {}  # {"path": str, "mtime": float|None, "rules": [(sources, targets), ...]}


def _load_field_matching_rules_from_knowledge() -> List[tuple[List[str], List[str]]]:
    """
    从 wanding_business_knowledge.md 的【字段匹配同义与规格】段落解析规则，
    用于字段匹配阶段同义扩展，提高命中率。返回 [(source_terms, target_terms), ...]。
    """
    global _FIELD_MATCHING_RULES_CACHE
    try:
        from inventory.config import config
        path_str = getattr(config, "WANDING_BUSINESS_KNOWLEDGE_PATH", None)
        if not path_str:
            return []
        p = Path(path_str)
        if not p.exists():
            return []
        mtime: Optional[float] = None
        try:
            mtime = p.stat().st_mtime
        except OSError:
            pass
        if (
            _FIELD_MATCHING_RULES_CACHE.get("path") == path_str
            and _FIELD_MATCHING_RULES_CACHE.get("mtime") == mtime
        ):
            return _FIELD_MATCHING_RULES_CACHE.get("rules") or []
        content = p.read_text(encoding="utf-8")
        rules: List[tuple[List[str], List[str]]] = []
        in_section = False
        for line in content.splitlines():
            line = line.strip()
            if "【字段匹配" in line or "【字段匹配同义" in line:
                in_section = True
                continue
            if in_section and line.startswith("【"):
                break
            if not in_section:
                continue
            # 解析 "- 源词 源词 → 检索词 检索词" 或 "  - ... → ..."
            if line.startswith("-"):
                line = line.lstrip("-").strip()
            if "→" in line:
                left, _, right = line.partition("→")
            elif "->" in line:
                left, _, right = line.partition("->")
            else:
                continue
            sources = [t.strip() for t in left.split() if t.strip()]
            targets = [t.strip() for t in right.split() if t.strip()]
            if sources and targets:
                rules.append((sources, targets))
        _FIELD_MATCHING_RULES_CACHE = {"path": path_str, "mtime": mtime, "rules": rules}
        return rules
    except Exception as e:
        logger.debug("加载字段匹配规则失败: %s", e)
        return []


def _apply_knowledge_expansion(keywords: str) -> str:
    """
    根据业务知识【字段匹配同义与规格】规则，在字段匹配前扩展询价词，
    使口语/同义词能命中库内品名（如 直接→直通 排水、热熔器→焊接机）。
    """
    if not (keywords or "").strip():
        return keywords
    rules = _load_field_matching_rules_from_knowledge()
    added: List[str] = []
    kw_lower = (keywords or "").lower()
    for sources, targets in rules:
        for src in sources:
            if src.lower() in kw_lower or re.search(re.escape(src), keywords, re.I):
                added.extend(targets)
                break
    if not added:
        return keywords.strip()
    return (keywords.strip() + " " + " ".join(added)).strip()


# PN ↔ MPa 双向扩展
# PN 是公称压力等级：PN10 = 1.0MPa, PN16 = 1.6MPa, PN12.5 = 1.25MPa（数字 × 0.1 ≈ MPa）
# PN 前需要词边界\b；MPa 前数字后不需要严格\b（因为数字前可能是空格/汉字/标点）
_PN_RE = re.compile(r'\bPN\s*(\d+(?:\.\d+)?)(?![\da-zA-Z_])', re.IGNORECASE)
# 数字+MPa，尾部用负前瞻排除纯 ASCII 字母数字后缀（如 MPaA、MPa1），汉字/空格/结尾均放行
_MPA_RE = re.compile(r'(\d+(?:\.\d+)?)\s*MPA(?![\da-zA-Z])', re.IGNORECASE)


def _format_pressure_value(value: float) -> str:
    """格式化压力值，去除浮点尾巴（如 1.25 而非 1.2500000001）"""
    formatted = f"{value:.2f}".rstrip('0').rstrip('.')
    return formatted


def _apply_pressure_expansion(keywords: str) -> str:
    """
    双向扩展 PN ↔ MPa，并做数值格式化。
    - PN -> MPa：PN16 -> 1.6MPa
    - MPa -> PN：1.25MPa -> PN12.5
    - 格式化：1.60MPa / 1.6 MPa / 1.6mpa -> 统一 1.6MPa
    - 去重：已扩展过的等价形式不重复追加
    """
    if not (keywords or "").strip():
        return ""

    # 预扫描已存在的 PN/MPa 数值，避免重复追加
    seen_pn: set[str] = set()
    seen_mpa: set[str] = set()

    def _scan_pn(m: re.Match) -> str:
        pn_val = float(m.group(1))
        seen_pn.add(_format_pressure_value(pn_val))
        return m.group(0)

    def _scan_mpa(m: re.Match) -> str:
        mpa_val = float(m.group(1))
        seen_mpa.add(_format_pressure_value(mpa_val))
        return m.group(0)

    _PN_RE.sub(_scan_pn, keywords)
    _MPA_RE.sub(_scan_mpa, keywords)

    additions: list[str] = []

    # PN -> MPa：仅当原词中不存在等价 MPa 值
    def _sub_pn_to_mpa(m: re.Match) -> str:
        pn_val = float(m.group(1))
        mpa_val = pn_val * 0.1
        mpa_formatted = _format_pressure_value(mpa_val)
        if mpa_formatted in seen_mpa:
            return m.group(0)
        seen_mpa.add(mpa_formatted)
        additions.append(f"{mpa_formatted}MPa")
        return m.group(0)

    # MPa -> PN：仅当原词中不存在等价 PN 值
    def _sub_mpa_to_pn(m: re.Match) -> str:
        mpa_val = float(m.group(1))
        pn_val = mpa_val * 10
        pn_formatted = _format_pressure_value(pn_val)
        if pn_formatted in seen_pn:
            return m.group(0)
        seen_pn.add(pn_formatted)
        additions.append(f"PN{pn_formatted}")
        return m.group(0)

    result = keywords
    result = _PN_RE.sub(_sub_pn_to_mpa, result)
    result = _MPA_RE.sub(_sub_mpa_to_pn, result)

    if additions:
        result = result + " " + " ".join(additions)

    return result


MM_TO_INCH = {
    "16": '1/2"', "20": '3/4"', "25": '1"', "32": '1-1/4"', "40": '1-1/2"',
    "50": '2"', "65": '2-1/2"', "75": '3"', "100": '4"', "125": '5"',
    "150": '6"', "200": '8"', "250": '10"', "300": '12"',
}
INCH_TO_MM = {v: k for k, v in MM_TO_INCH.items()}

# 国内口语「N 分」管径 → 公称 DN 数字（与价格库 Describrition 中 DN15/DN20 对齐）
# 4 分 ≈ 1/2" → DN15；6 分 ≈ 3/4" → DN20（避免把「4」单拆出来导致与 DN15 无法匹配）
FEN_TO_MM_STR = {
    "4": "15",
    "6": "20",
}
# 整数「N 寸」→ DN（常用对照，与 _split_tokens 中 \d+寸 整段提取配合）
CUN_INTEGER_TO_MM_STR = {
    "1": "25",
    "2": "50",
    "3": "80",
    "4": "100",
    "5": "125",
    "6": "150",
    "8": "200",
    "10": "250",
    "12": "300",
}

# 脚本/旧测试兼容别名（仅「寸」映射；「分」见 FEN_TO_MM_STR）
CUN_TO_MM = CUN_INTEGER_TO_MM_STR

# 日标 PVC-U 排水管件：外径口语数字 → 公称通径 DN
# 这些数字不在 MM_TO_INCH（标准 DN 系列）中，出现在查询里必然是外径，不是 DN：
#   OD63  → DN50 (2")   OD110 → DN100 (4")   OD160 → DN150 (6")
# 添加 DN 等价后，日标产品（Describrition 含 DN100/DN150）才能进入候选；
# 原 OD 值同时保留，让国标产品（Describrition 含 dn110/dn160）也能命中，
# 最终由 LLM 或 Product_Type 过滤在两类候选中选出正确结果。
OD_TO_DN_JIS: dict[str, str] = {
    "63": "50",    # OD63  → DN50  (2")
    "110": "100",  # OD110 → DN100 (4")
    "160": "150",  # OD160 → DN150 (6")
}


def _normalize(s: str) -> str:
    s = (s or "").lower().strip()
    s = s.replace("－", "-").replace("—", "-").replace("（", "(").replace("）", ")")
    s = re.sub(r"[_\t]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _normalize_chinese_number_order(s: str) -> str:
    """
    Keep legacy token normalization behavior for tests and callers:
    reorder "50三通" -> "三通50" while preserving more specific phrases.
    """
    text = s or ""
    m = re.fullmatch(r"(\d+)([\u4e00-\u9fff]{2})", text)
    if m:
        return f"{m.group(2)}{m.group(1)}"
    return text


def _get_synonym_words(word: str) -> frozenset:
    return _SYNONYM_TO_GROUP.get(word, frozenset({word}))


def _expand_unit_tokens(token: str, material: Optional[str] = None) -> set:
    eqs = {token}
    # 支持分数英寸规格（如 3/4、1-1/4），统一补全为带引号形式参与映射
    if re.fullmatch(r"\d+(?:-\d+)?/\d+", token):
        token = token + '"'
        eqs.add(token)
    # 行业场景下将 DE 视为 DN 的等价写法（如 DE50≈DN50）
    if token.startswith("de"):
        num = token[2:]
        if not num.isdigit():
            return eqs
        if num in MM_TO_INCH:
            eqs.add(MM_TO_INCH[num])
        eqs.add(num)
        eqs.add("dn" + num)
        return eqs
    if token.startswith("dn"):
        num = token[2:]
        if num in MM_TO_INCH:
            eqs.add(MM_TO_INCH[num])
        eqs.add(num)
        return eqs
    if token.isdigit() and token in MM_TO_INCH:
        eqs.add("dn" + token)
        eqs.add(MM_TO_INCH[token])
        return eqs
    # 日标外径 → DN 等价（63→50, 110→100, 160→150）
    # 原 OD 值（token）已在 eqs 中，保留以命中国标产品；DN 等价用于命中日标产品
    if token.isdigit() and token in OD_TO_DN_JIS:
        dn_num = OD_TO_DN_JIS[token]
        eqs.add(dn_num)
        eqs.add("dn" + dn_num)
        if dn_num in MM_TO_INCH:
            eqs.add(MM_TO_INCH[dn_num])
        return eqs
    if token in INCH_TO_MM:
        eqs.add(INCH_TO_MM[token])
        eqs.add("dn" + INCH_TO_MM[token])
        return eqs
    m_fen = re.fullmatch(r"(\d+)分", token)
    if m_fen:
        n = m_fen.group(1)
        if n in FEN_TO_MM_STR:
            mm = FEN_TO_MM_STR[n]
            eqs.add("dn" + mm)
            eqs.add(mm)
            if mm in MM_TO_INCH:
                eqs.add(MM_TO_INCH[mm])
        return eqs
    m_cun = re.fullmatch(r"(\d+)寸", token)
    if m_cun:
        n = m_cun.group(1)
        if n in CUN_INTEGER_TO_MM_STR:
            mm = CUN_INTEGER_TO_MM_STR[n]
            eqs.add("dn" + mm)
            eqs.add(mm)
            if mm in MM_TO_INCH:
                eqs.add(MM_TO_INCH[mm])
        return eqs
    return eqs


def _expand_token_with_synonyms_and_units(token: str, material: Optional[str] = None) -> set:
    synonyms = _get_synonym_words(token)
    expanded: set = set()
    for syn in synonyms:
        expanded |= _expand_unit_tokens(syn, material=material)
    return expanded


def _is_inch_token(token: str) -> bool:
    """是否为英寸规格 token（如 3/4"、1-1/4"、4"）。"""
    t = (token or "").strip()
    return bool(
        re.fullmatch(r"\d+(?:-\d+)?/\d+\"", t)
        or re.fullmatch(r"\d+(?:\.\d+)?\"", t)
    )


def _should_apply_inch_exact_priority(
    query_size_tokens: set[str],
    query_inch_tokens: set[str],
) -> bool:
    """
    「英寸精确命中优先」剪切：仅当查询里的英寸与 DN 数字在标准对照上一致时才启用。

    若用户混写互斥的公称径与英寸（例如 DN20 对应 3/4\"，却写了 1/2\"≈DN16），
    则不做剪切，否则仅写 dn20、库内无字面 1/2\" 的正确行会被误杀，而带 3/4\"x1/2\" 的管件会留下。
    """
    if not query_inch_tokens:
        return False
    dn_numbers: set[str] = set()
    for t in query_size_tokens:
        tl = (t or "").lower().strip()
        if tl.startswith("dn") and len(tl) > 2 and tl[2:].isdigit():
            dn_numbers.add(tl[2:])
    if not dn_numbers:
        return True
    for inch_t in query_inch_tokens:
        mm = INCH_TO_MM.get(inch_t)
        if mm is not None and mm not in dn_numbers:
            return False
    return True


def _split_tokens(text: str) -> List[str]:
    text = _normalize(text)
    tokens: List[str] = []
    # 先提取英寸分数规格，避免被后续纯数字提取拆成 3、4 这类噪声 token
    # 例：3/4"、1-1/4"
    for m in re.finditer(r"\d+(?:\s*-\s*\d+)?\s*/\s*\d+\s*[\"”″]?", text):
        raw = m.group()
        compact = re.sub(r"\s+", "", raw)
        compact = compact.replace("”", '"').replace("″", '"')
        if "/" in compact:
            if not compact.endswith('"'):
                compact = compact + '"'
            tokens.append(compact)
    text = re.sub(r"\d+(?:\s*-\s*\d+)?\s*/\s*\d+\s*[\"”″]?", " ", text)
    for m in re.finditer(r"\d+(?:\.\d+)?\s*[\"”″]", text):
        raw = m.group()
        compact = re.sub(r"\s+", "", raw)
        compact = compact.replace("”", '"').replace("″", '"')
        if not compact.endswith('"'):
            compact = compact + '"'
        tokens.append(compact)
    text = re.sub(r"\d+(?:\.\d+)?\s*[\"”″]", " ", text)
    # 「N分」口语（4分=DN15、6分=DN20）：整段提取，避免拆成孤立数字「4」导致规格过滤失败
    for m in re.finditer(r"\d+\s*分", text):
        tokens.append(re.sub(r"\s+", "", m.group()))
    text = re.sub(r"\d+\s*分", " ", text)
    # 「N寸」整数寸（如 2寸→DN50）
    for m in re.finditer(r"\d+\s*寸", text):
        tokens.append(re.sub(r"\s+", "", m.group()))
    text = re.sub(r"\d+\s*寸", " ", text)
    for m in re.finditer(r"dn\s*(\d+)", text, re.I):
        tokens.append("dn" + m.group(1))
        tokens.append(m.group(1))
    text = re.sub(r"dn\s*\d+", " ", text, flags=re.I)
    for m in re.finditer(r"\d+(?:\.\d+)?", text):
        tokens.append(m.group())
    text = re.sub(r"\d+(?:\.\d+)?", " ", text)
    for m in re.finditer(r"[\u4e00-\u9fff]+", text):
        tok = m.group()
        tokens.append(tok)
        # 长中文整段（如「度弯头带检查口」）在品名中常被括号/符号隔开，拆成单字以便「度」→「°」、「弯头」「带」「检查口」等能分别命中
        if len(tok) > 2:
            for c in tok:
                tokens.append(c)
    return list(dict.fromkeys(tokens))


def _expand_keyword_with_synonyms(keyword: str) -> List[str]:
    queries: set = {keyword}
    for syn in _SORTED_SYNONYMS:
        new_queries: set = set()
        for q in queries:
            if syn in q:
                for replacement in _SYNONYM_TO_GROUP[syn]:
                    new_queries.add(q.replace(syn, replacement))
        if new_queries:
            queries.update(new_queries)
    return list(queries)


def _safe_to_float(val: Any) -> Optional[float]:
    """将单元格值安全转为 float；非法或空值返回 None。"""
    if val is None:
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    return f


def _normalize_code_for_match(val: Any) -> str:
    """物料编号用于匹配时的规范化：数字去掉尾部的 .0，避免 8010072480.0 与 8010072480 不匹配。"""
    if val is None or (hasattr(pd, "isna") and pd.isna(val)):
        return ""
    s = str(val).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except (TypeError, ValueError):
        return s


def normalize_price(raw_price: Any) -> float:
    """
    规范化用户输入的价格为 float（人民币金额）。

    规则（auto_fix 风格）：
    - 若本身是数字（int/float），直接返回 float 值。
    - 若是字符串：
      - 去除首尾空白、常见货币符号（¥/￥/元/RMB 等）及内部空格。
      - 将全角逗号替换为半角，移除千分位逗号。
      - 若存在多个小数点：保留最后一个点作为小数点，其余点视作分隔符移除（例如 "7.858.0" → "7858.0"）。
      - 清洗后字符串必须形如 `^[+-]?\\d+(\\.\\d+)?$`，否则视为格式不合法。
    - 解析失败或结果为 NaN/无穷大时抛出 ValueError。
    """
    # 已经是数字的情况，直接信任调用方
    if isinstance(raw_price, (int, float)) and not isinstance(raw_price, bool):
        value = float(raw_price)
        if value != value or value in (float("inf"), float("-inf")):
            raise ValueError(f"价格数值非法: {raw_price!r}")
        return value

    s = str(raw_price).strip()
    if not s:
        raise ValueError("价格不能为空。")

    # 去掉常见货币符号与全角变体
    for sym in ("¥", "￥", "元", "RMB", "rmb"):
        s = s.replace(sym, "")

    # 统一全角/半角逗号与点
    s = s.replace("，", ",").replace("．", ".")
    # 去掉空格
    s = s.replace(" ", "")

    # 先移除千分位逗号
    s = s.replace(",", "")

    # 处理多个小数点：保留最后一个，其余删除
    if s.count(".") > 1:
        last_dot = s.rfind(".")
        cleaned_chars = []
        for idx, ch in enumerate(s):
            if ch == "." and idx != last_dot:
                continue
            cleaned_chars.append(ch)
        s = "".join(cleaned_chars)

    # 允许前导正负号，其余必须是数字或至多一个小数点
    import re as _re

    if not _re.fullmatch(r"[+-]?\d+(\.\d+)?", s):
        raise ValueError(f"价格格式不合法: {raw_price!r}")

    try:
        value = float(s)
    except (TypeError, ValueError):
        raise ValueError(f"价格无法解析: {raw_price!r}")

    if value != value or value in (float("inf"), float("-inf")):
        raise ValueError(f"价格数值非法: {raw_price!r}")
    return value


def search_fuzzy(
    df: pd.DataFrame,
    keyword: str,
    field: str = "Describrition",
) -> List[tuple[dict[str, Any], float]]:
    """
    DataBase-style fuzzy search.
    Returns [(row_dict, score), ...] sorted by score desc.
    row_dict: {code, matched_name, unit_price}

    优化点：
    - 使用 load_wanding_df 预计算的 norm_text / spec_tokens 列，消除每行的 regex 开销
    - q_eq 提出行循环，每次查询只计算一次
    - set 交集（q_eq & product_specs）替代 any(eq in set for eq in set)
    - 单字 token 权重 _SINGLE_CHAR_WEIGHT（0.5），避免单字命中过度拉高得分
    """
    results: dict = {}
    has_precomputed = "norm_text" in df.columns and "spec_tokens" in df.columns

    for kw in _expand_keyword_with_synonyms(keyword.strip()):
        norm_kw = _normalize(kw)
        chinese_tokens = _split_tokens(norm_kw)
        material_tokens = re.findall(r"pvc|ppr|pe|hdpe", norm_kw)
        query_size_tokens = {t for t in chinese_tokens if re.search(r"\d", t) and not t.endswith("°")}
        query_inch_tokens = {t for t in query_size_tokens if _is_inch_token(t)}
        query_text_tokens = {
            t for t in chinese_tokens if not (re.search(r"\d", t) and not t.endswith("°"))
        }
        query_material = material_tokens[0] if material_tokens else None

        # q_eq 提出行循环：只依赖 q_spec + query_material，与当前行无关
        spec_equivs: dict[str, frozenset] = {
            q_spec: _expand_token_with_synonyms_and_units(q_spec, material=query_material)
            for q_spec in query_size_tokens
        }

        # 按单字/多字分类，单字在分母中按 _SINGLE_CHAR_WEIGHT 计入
        multi_text = {t for t in query_text_tokens if len(t) > 1}
        single_text = {t for t in query_text_tokens if len(t) == 1}
        total_weight = (
            len(query_size_tokens)
            + len(multi_text)
            + len(single_text) * _SINGLE_CHAR_WEIGHT
        )

        iter_rows: list[tuple[dict[str, Any], float, int]] = []
        for row in df.itertuples(index=False):
            row_id = getattr(row, "Material", getattr(row, "Describrition", str(row)))
            raw_text = str(getattr(row, field, ""))

            # 使用预计算列，fallback 到实时计算（兼容未预计算的 df）
            if has_precomputed:
                normalized_text: str = row.norm_text
                product_specs: frozenset = row.spec_tokens
            else:
                normalized_text = _normalize(raw_text)
                product_specs = frozenset(
                    t for t in _split_tokens(raw_text) if re.search(r"\d", t)
                )

            if material_tokens and not all(m.lower() in normalized_text for m in material_tokens):
                continue

            # set 交集替代 any(eq in product_specs for eq in q_eq)
            size_hits = sum(1 for q_eq in spec_equivs.values() if q_eq & product_specs)
            if query_size_tokens and size_hits == 0:
                continue
            inch_exact_hits = sum(1 for t in query_inch_tokens if t in product_specs)

            # 多字命中（权重 1.0）+ 单字命中（权重 _SINGLE_CHAR_WEIGHT）
            def _text_match(t: str) -> bool:
                return t.lower() in normalized_text or (t == "度" and "°" in normalized_text)

            multi_hits = sum(1 for t in multi_text if _text_match(t))
            single_hits = sum(1 for t in single_text if _text_match(t))
            # 过滤用原始命中数（单字也算），得分用加权值
            if query_text_tokens and (multi_hits + single_hits) == 0:
                continue

            hit_weight = size_hits + multi_hits + single_hits * _SINGLE_CHAR_WEIGHT
            score = hit_weight / total_weight if total_weight > 0 else 0.0

            if score > 0 and (row_id not in results or score > results[row_id][1]):
                row_dict: dict[str, Any] = {
                    "code": str(getattr(row, "Material", "")),
                    "matched_name": raw_text,
                }
                if hasattr(row, "unit_price"):
                    row_dict["unit_price"] = getattr(row, "unit_price", 0.0)
                if hasattr(row, "Product_Type"):
                    row_dict["Product_Type"] = str(getattr(row, "Product_Type", "") or "").strip()
                iter_rows.append((row_dict, score, inch_exact_hits))

        # 英寸优先：若查询里显式给了英寸，且存在英寸精确命中的候选，
        # 则只保留英寸精确命中，避免被 dn 等价扩展引入跨体系误匹配。
        # （当 DN 与英寸互斥时跳过，见 _should_apply_inch_exact_priority）
        if (
            _should_apply_inch_exact_priority(query_size_tokens, query_inch_tokens)
            and query_inch_tokens
            and any(inch_hits > 0 for _, _, inch_hits in iter_rows)
        ):
            iter_rows = [r for r in iter_rows if r[2] > 0]

        for row_dict, score, _inch_hits in iter_rows:
            row_id = row_dict.get("code") or row_dict.get("matched_name")
            if row_id not in results or score > results[row_id][1]:
                results[row_id] = (row_dict, score)

    out = list(results.values())
    out.sort(key=lambda x: x[1], reverse=True)
    return out


def _load_one_sheet(ws, price_col: int) -> list[dict]:
    """从已打开的 worksheet 读出一张表的行（Material, Describrition, Describrition_English, Product_Type, unit_price）。需覆盖 E 档列(0-based 18)，故 max_col=20。"""
    rows = []
    for row in ws.iter_rows(max_col=20):
        cells = [getattr(c, "value", None) for c in row]
        if len(cells) > 2 and cells[2]:
            up = 0.0
            if len(cells) > price_col and cells[price_col] is not None:
                try:
                    up = float(cells[price_col])
                except (ValueError, TypeError):
                    pass
            rows.append({
                "Material": str(cells[1] or "").strip(),
                "Describrition": str(cells[2] or "").strip(),
                "Describrition_English": str(cells[3] or "").strip() if len(cells) > 3 else "",
                "Product_Type": str(cells[4] or "").strip() if len(cells) > 4 else "",
                "unit_price": up,
            })
    return rows


def load_wanding_df(
    path: str | Path,
    sheet_name: str = "管材",
    customer_level: str = "B",
) -> pd.DataFrame:
    """Load 万鼎 price library as DataFrame. 默认加载「管材」+「国标管件」两个 sheet 并合并，以便匹配带检查口弯头、管帽等国标管件。"""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl 未安装，万鼎模糊匹配不可用")
        return pd.DataFrame()

    p = Path(path)
    if p.is_absolute() and p.exists():
        pass
    elif not p.is_absolute():
        root = Path(__file__).resolve().parent.parent.parent
        p = root / p
    if not p.exists():
        logger.warning("万鼎价格库不存在: %s", p)
        return pd.DataFrame()

    level = _normalize_price_level(customer_level)
    price_col = PRICE_COLS.get(level, PRICE_COLS["B"])

    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        all_rows: list[dict] = []
        # 先加载管材
        ws_guan = wb["管材"] if "管材" in wb.sheetnames else (wb.active or wb[wb.sheetnames[0]])
        all_rows.extend(_load_one_sheet(ws_guan, price_col))
        # 若存在国标管件 sheet，一并加载（8020020643 带检查口弯头、8020020205 管帽等在此表）
        if "国标管件" in wb.sheetnames:
            all_rows.extend(_load_one_sheet(wb["国标管件"], price_col))
        wb.close()
        df = pd.DataFrame(all_rows)
        if not df.empty:
            # 预计算 normalized text 和规格 token 集，避免 search_fuzzy 每行重算
            df["norm_text"] = df["Describrition"].apply(_normalize)
            df["spec_tokens"] = df["Describrition"].apply(
                lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
            )
        return df
    except Exception as e:
        logger.warning("加载万鼎价格库失败: %s", e)
        return pd.DataFrame()


def invalidate_wanding_cache() -> None:
    """清除万鼎 DataFrame 缓存（admin 更新 Neon 数据后调用）。"""
    global _full_df_cache
    with _df_cache_lock:
        _df_cache.clear()
    with _full_df_lock:
        _full_df_cache = None
    logger.info("wanding_fuzzy_matcher: DataFrame caches cleared")


def _level_to_db_price_field(level: str) -> Optional[str]:
    """A/B/C/D 档 → DB 列名；出厂/采购/E 档等返回 None（走本地 xlsx）。"""
    lu = level.upper()
    if "FACTORY" in lu or "PURCHASE" in lu:
        return None
    if lu.startswith("E_") or lu == "E":
        return None
    if lu.startswith("A_") or lu == "A":
        return "price_a"
    if lu.startswith("B_") or lu == "B":
        return "price_b"
    if lu.startswith("C_") or lu == "C":
        return "price_c"
    if lu.startswith("D_") or lu in ("D", "D_LOW", "D_NOADJ", "D_WHOLESALE"):
        return "price_d"
    return None


def _find_col(columns: list[dict], *keywords: str) -> str | None:
    """从 columns JSON 找第一个 name（lower）同时包含所有 keyword（lower）的列。"""
    for col in columns:
        name = (col.get("name") or "").lower()
        if all(kw.lower() in name for kw in keywords):
            return col["name"]
    return None


def _try_load_from_custom_library(level: str) -> Optional[pd.DataFrame]:
    """当 price_library 固定表为空时，从名含"万鼎"/"价格库"的自定义库拉数据。"""
    try:
        from admin import repository
        from inventory.config import config

        libs = repository.list_libraries()
        patterns = config.PRICE_LIB_NAME_PATTERNS
        matched = [lib for lib in libs if any(p in (lib.get("name") or "") for p in patterns)]
        if not matched:
            return None

        lib = max(matched, key=lambda x: x["id"])
        table_name = lib.get("table_name") or ""
        columns = lib.get("columns") or []
        if not table_name or not columns:
            return None

        col_material = _find_col(columns, config.PRICE_LIB_COL_MATERIAL_KW)
        col_desc = _find_col(columns, config.PRICE_LIB_COL_DESC_KW)
        field = _level_to_db_price_field(level)
        level_kw_map = {
            "price_a": config.PRICE_LIB_COL_PRICE_A_KW,
            "price_b": config.PRICE_LIB_COL_PRICE_B_KW,
            "price_c": config.PRICE_LIB_COL_PRICE_C_KW,
            "price_d": config.PRICE_LIB_COL_PRICE_D_KW,
        }
        price_kw = level_kw_map.get(field or "")
        col_price = _find_col(columns, *price_kw) if price_kw else None

        if not col_desc:
            logger.warning("自定义价格库找不到描述列 (keywords=%s)", config.PRICE_LIB_COL_DESC_KW)
            return None

        rows = repository.fetch_all_library_rows(table_name)
        if not rows:
            return None

        records = []
        for r in rows:
            material = str(r.get(col_material) or "" if col_material else "").strip()
            desc = str(r.get(col_desc) or "").strip()
            if not desc:
                continue
            up_f = 0.0
            if col_price:
                try:
                    up_f = float(r.get(col_price) or 0)
                except (TypeError, ValueError):
                    up_f = 0.0
            records.append({"Material": material, "Describrition": desc, "unit_price": up_f})

        if not records:
            return None
        df = pd.DataFrame(records)
        df["norm_text"] = df["Describrition"].apply(_normalize)
        df["spec_tokens"] = df["Describrition"].apply(
            lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
        )
        logger.info(
            "wanding_fuzzy_matcher: loaded %d rows from custom library '%s' (level=%s)",
            len(df), lib.get("name"), level,
        )
        return df
    except Exception as e:
        logger.warning("_try_load_from_custom_library 失败: %s", e)
        return None


def _try_load_from_db(level: str) -> Optional[pd.DataFrame]:
    """从 Neon admin 缓存行构建 DataFrame；无数据或不可用则返回 None（fallback xlsx）。"""
    try:
        from admin.cache import get_price_library_rows

        rows = get_price_library_rows()
        if not rows:
            return _try_load_from_custom_library(level)
        field = _level_to_db_price_field(level)
        if field is None:
            return None
        records = []
        for r in rows:
            up = r.get(field)
            try:
                up_f = float(up) if up is not None else 0.0
            except (TypeError, ValueError):
                up_f = 0.0
            records.append(
                {
                    "Material": str(r.get("material") or "").strip(),
                    "Describrition": str(r.get("description") or "").strip(),
                    "Describrition_English": str(r.get("description_english") or "").strip(),
                    "Product_Type": str(r.get("product_type") or "").strip(),
                    "unit_price": up_f,
                }
            )
        df = pd.DataFrame(records)
        if df.empty:
            return None
        df["norm_text"] = df["Describrition"].apply(_normalize)
        df["spec_tokens"] = df["Describrition"].apply(
            lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
        )
        logger.info("wanding_fuzzy_matcher: loaded %d rows from DB (level=%s)", len(df), level)
        return df
    except Exception as e:
        logger.warning("_try_load_from_db 失败，将 fallback 读 xlsx: %s", e)
        return None


# 缓存 DataFrame，按 path:level 隔离
_df_cache: dict[str, pd.DataFrame] = {}
_df_cache_lock = threading.Lock()


def _get_cached_df(path, customer_level: str) -> pd.DataFrame:
    """线程安全地获取 DataFrame。优先 Neon（有数据时），否则读本地 xlsx。"""
    level = _normalize_price_level(customer_level)
    cache_key = f"{path}:{level}"
    with _df_cache_lock:
        if cache_key not in _df_cache:
            df_db = _try_load_from_db(level)
            if df_db is not None and not df_db.empty:
                _df_cache[cache_key] = df_db
            else:
                _df_cache[cache_key] = load_wanding_df(path, customer_level=level)
        return _df_cache[cache_key]


# --------- 利润率查询（按 code / 完整名称 + 价格）---------

_full_df_cache: Optional[pd.DataFrame] = None
_full_df_lock = threading.Lock()


def _load_full_price_df(path: str | Path) -> pd.DataFrame:
    """加载完整万鼎价格库 DataFrame（包含所有价格与利润率列），供利润率查询使用。"""
    global _full_df_cache
    if _full_df_cache is not None:
        return _full_df_cache
    with _full_df_lock:
        if _full_df_cache is not None:
            return _full_df_cache
        p = Path(path)
        if not p.is_absolute():
            root = Path(__file__).resolve().parent.parent.parent
            p = root / p
        if not p.exists():
            logger.warning("万鼎价格库不存在: %s", p)
            _full_df_cache = pd.DataFrame()
            return _full_df_cache
        try:
            # 同时加载「管材」与「国标管件」两个 sheet 并合并
            sheets = pd.read_excel(p, sheet_name=None)
            frames: list[pd.DataFrame] = []
            for name, df in sheets.items():
                if name in ("管材", "国标管件") or not frames:
                    frames.append(df)
            df_all = pd.concat(frames, ignore_index=True)
            _full_df_cache = df_all
            return _full_df_cache
        except Exception as e:
            logger.warning("加载完整万鼎价格库失败: %s", e)
            _full_df_cache = pd.DataFrame()
            return _full_df_cache


def _compute_profit_for_price(row: pd.Series, price: float) -> dict[str, Any]:
    """给定一行价格库记录与用户价，计算匹配档位及其利润率，并返回所有档位价格+利润率。

    行为：
    - 仅当某档位价格与给定 price 在容差内相等（绝对误差 ≤ _tolerance，_tolerance = max(0.01, |target|×1e-5)）时，
      才设置 matched_* 字段，避免浮点/四舍五入导致“有记录但匹配不到档位”。
    - 若没有任何档位价格在容差内与 price 相等，则 matched_price_level/matched_price/matched_profit 保持为 None，
      不再回退到“距离最小”的近似档位。
    """
    all_levels: list[dict[str, Any]] = []
    for level in ("A_QUOTE", "B_QUOTE", "C_QUOTE", "D_QUOTE", "D_LOW", "E_QUOTE"):
        price_col = PRICE_COLS.get(level)
        profit_col = PROFIT_COLS.get(level)
        if price_col is None:
            continue
        price_val = None
        if len(row) > price_col:
            price_val = _safe_to_float(row.iloc[price_col])
        if price_val is None or price_val == 0:
            continue
        profit_val = None
        if profit_col is not None and len(row) > profit_col:
            profit_val = _safe_to_float(row.iloc[profit_col])
        all_levels.append(
            {
                "level": level,
                "price": price_val,
                "profit": profit_val,
                "level_display": PRICE_LEVEL_DISPLAY_NAMES.get(level, level),
            }
        )
    matched_level = None
    matched_price = None
    matched_profit = None
    if all_levels:
        target = float(price)
        # 档位匹配：允许极小误差（浮点/四舍五入），避免库中 21810.0 与用户 21810 或 21809.99 判为不匹配
        _tolerance = max(0.01, abs(target) * 1e-5)
        exact = [entry for entry in all_levels if abs(entry["price"] - target) <= _tolerance]
        if exact:
            best = exact[0]
            matched_level = best["level"]
            matched_price = best["price"]
            matched_profit = best["profit"]
    return {
        "code": str(row.get("Material") or row.get("code") or "").strip(),
        "name": str(row.get("Describrition") or row.get("Description") or "").strip(),
        "matched_price_level": matched_level,
        "matched_price": matched_price,
        "matched_profit": matched_profit,
        "all_levels": all_levels,
    }


def get_profit_rows_by_code(code: str, price: float, path: str | Path) -> list[dict[str, Any]]:
    """按 Material code 精确过滤价格库，并为每行计算与给定价格对应的利润率。"""
    df = _load_full_price_df(path)
    if df.empty:
        return []
    code_norm = _normalize_code_for_match(code)
    if "Material" in df.columns:
        # 统一规范化：库中可能是 8010072480.0，用户传 8010072480，需一致
        mask = df["Material"].apply(lambda v: _normalize_code_for_match(v) == code_norm)
        rows = df[mask]
    else:
        rows = pd.DataFrame()
    return [_compute_profit_for_price(row, price) for _, row in rows.iterrows()]


def get_profit_rows_by_name(name: str, price: float, path: str | Path) -> list[dict[str, Any]]:
    """按完整中文名称过滤价格库，并为每行计算与给定价格对应的利润率。"""
    df = _load_full_price_df(path)
    if df.empty:
        return []
    name_norm = _normalize(name)
    col = "Describrition" if "Describrition" in df.columns else "Description"
    series = df[col].astype(str).apply(_normalize)
    rows = df[series == name_norm]
    return [_compute_profit_for_price(row, price) for _, row in rows.iterrows()]


def match_fuzzy(
    keywords: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
    product_type: Optional[str] = None,
) -> Optional[dict[str, Any]]:
    """
    DataBase-style 模糊匹配，返回最佳单结果。
    返回 {code, matched_name, unit_price} 或 None。
    先按业务知识【字段匹配补充规则】扩展检索词，再做同义/外语替换（SYNONYM_GROUPS、QUERY_TERM_TO_CHINESE）。
    """
    keywords = (keywords or "").strip()
    keywords = _apply_knowledge_expansion(keywords)
    keywords = _apply_pressure_expansion(keywords)
    keywords = _normalize_keyword_terms(keywords)
    keywords = _strip_query_intent_terms(keywords)
    if not keywords:
        return None

    from inventory.config import config
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level)
    if df.empty:
        return None
    if product_type:
        if "Product_Type" not in df.columns:
            logger.warning("Product_Type 严格过滤失败：数据源缺少 Product_Type 列，filter=%s", product_type)
            return None
        df = df[df["Product_Type"].astype(str).str.strip() == product_type]
        if df.empty:
            logger.warning("Product_Type 过滤后无候选: %s", product_type)
            return None

    results = search_fuzzy(df, keywords)
    if not results:
        return None

    row_dict, _ = results[0]
    return {
        "code": (row_dict.get("code") or "").strip(),
        "matched_name": (row_dict.get("matched_name") or "")[:200],
        "unit_price": float(row_dict.get("unit_price", 0) or 0),
    }


def match_fuzzy_candidates(
    keywords: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
    max_candidates: int = 20,
    max_score_tiers: Optional[int] = None,
    min_score: Optional[float] = None,
    min_score_gap: Optional[float] = None,
    product_type: Optional[str] = None,
) -> List[dict[str, Any]]:
    """
    返回候选列表，每项含 code, matched_name, unit_price, score。
    - max_score_tiers 为 None：按分数排序取前 max_candidates 条。
    - max_score_tiers 为 N（如 2）：取分数前 N 档，每档全部返回（如 top1 有 3 条、top2 有 2 条则共 5 条）。
    - min_score：若最高分低于该阈值，则视为整体未命中，返回空列表。
    - min_score_gap：若 top1 分数与 top2 之差 ≥ 该值，则仅保留 top1 作为唯一高置信度候选。
    先按业务知识【字段匹配补充规则】扩展检索词，再做同义/外语替换。
    """
    keywords = (keywords or "").strip()
    keywords = _apply_knowledge_expansion(keywords)
    keywords = _apply_pressure_expansion(keywords)
    keywords = _normalize_keyword_terms(keywords)
    keywords = _strip_query_intent_terms(keywords)
    if not keywords:
        return []

    from inventory.config import config
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level)
    if df.empty:
        return []
    if product_type:
        if "Product_Type" not in df.columns:
            logger.warning("Product_Type 严格过滤失败：数据源缺少 Product_Type 列，filter=%s", product_type)
            return []
        df = df[df["Product_Type"].astype(str).str.strip() == product_type]
        if df.empty:
            logger.warning("Product_Type 过滤后无候选: %s", product_type)
            return []

    results = search_fuzzy(df, keywords)
    if not results:
        return []

    # 默认阈值从配置读取（如未显式传入）
    if min_score is None:
        min_score = getattr(config, "INVENTORY_MIN_SCORE", None)
    if min_score_gap is None:
        min_score_gap = getattr(config, "INVENTORY_MIN_SCORE_GAP", None)

    top_score = results[0][1]
    # 若最高分低于阈值，则视为整体未命中，直接返回空列表
    if isinstance(min_score, (int, float)) and top_score < float(min_score):
        return []

    # 若 top1 与 top2 分数差足够大，则视为唯一高置信度候选，直接截断为单候选
    if (
        isinstance(min_score_gap, (int, float))
        and len(results) >= 2
        and (top_score - results[1][1]) >= float(min_score_gap)
    ):
        results = results[:1]
    elif max_score_tiers is not None and max_score_tiers > 0:
        # 取前 max_score_tiers 个分数档，每档全部返回
        tiers: List[float] = []
        for _rd, score in results:
            if score not in tiers:
                tiers.append(score)
                if len(tiers) >= max_score_tiers:
                    break
        results = [(rd, s) for rd, s in results if s in tiers]
    else:
        results = results[:max_candidates]
    out = []
    for row_dict, score in results:
        out.append({
            "code": (row_dict.get("code") or "").strip(),
            "matched_name": (row_dict.get("matched_name") or "")[:200],
            "unit_price": float(row_dict.get("unit_price", 0) or 0),
            "score": round(score, 4),
            "Product_Type": (row_dict.get("Product_Type") or "").strip(),
        })
    return out


def match_english_candidates(
    keywords: str,
    customer_level: str = "B",
    price_library_path: str | Path | None = None,
    max_candidates: int = 20,
    product_type: Optional[str] = None,
) -> List[dict[str, Any]]:
    """
    英文 query → Describrition_English CONTAINS 匹配。
    将 keywords 按空白/标点拆分：长度 ≥2 的片段，以及单独的数字规格（如 3、50）作为 token；
    每行需全部 token 均出现在英文描述（小写）中才收录。
    返回 [{code, matched_name, description_english, unit_price, source}, ...]，最多 max_candidates 条。
    matched_name 为中文 Describrition，保持下游结构一致。
    """
    from inventory.config import config

    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level)
    if df.empty or "Describrition_English" not in df.columns:
        return []
    if product_type:
        if "Product_Type" not in df.columns:
            logger.warning("Product_Type 严格过滤失败：数据源缺少 Product_Type 列，filter=%s", product_type)
            return []
        df = df[df["Product_Type"].astype(str).str.strip() == product_type]
        if df.empty:
            logger.warning("Product_Type 过滤后无候选: %s", product_type)
            return []

    raw_tokens = re.split(r'[\s\"\'\-/\\]+', (keywords or "").strip())
    tokens: list[str] = []
    for t in raw_tokens:
        t = (t or "").strip()
        if not t:
            continue
        tl = t.lower()
        if len(tl) >= 2:
            tokens.append(tl)
        elif tl.isdigit():
            # 英寸/规格数字（如 3" pipe 中的 3）单独参与 CONTAINS
            tokens.append(tl)
    if not tokens:
        return []

    results: list[dict[str, Any]] = []
    for row in df.itertuples(index=False):
        en_desc = str(getattr(row, "Describrition_English", "") or "").lower()
        if not en_desc:
            continue
        if all(t in en_desc for t in tokens):
            results.append({
                "code": str(getattr(row, "Material", "")).strip(),
                "matched_name": str(getattr(row, "Describrition", "")).strip(),
                "description_english": str(getattr(row, "Describrition_English", "")).strip(),
                "Product_Type": str(getattr(row, "Product_Type", "")).strip(),
                "unit_price": float(getattr(row, "unit_price", 0) or 0),
                "source": "英文字段匹配",
            })
        if len(results) >= max_candidates:
            break
    return results


def get_wanding_price_by_code(
    code: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
) -> Optional[dict[str, Any]]:
    """
    按产品编号（Material）在万鼎价格表中精确查找，返回该 code 的单价及名称。
    用于历史匹配拿到 code 后，从万鼎表把价格补全。
    返回 {code, matched_name, unit_price} 或 None（万鼎表无此 code）。
    """
    code = (code or "").strip()
    if not code:
        return None
    from inventory.config import config
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level)
    if df.empty or "Material" not in df.columns:
        return None
    code_norm = _normalize_code_for_match(code)
    row = df[df["Material"].apply(lambda v: _normalize_code_for_match(v) == code_norm)]
    if row.empty:
        return None
    r = row.iloc[0]
    return {
        "code": code,
        "matched_name": str(r.get("Describrition", "") or "")[:200],
        "unit_price": float(r.get("unit_price", 0) or 0),
    }
