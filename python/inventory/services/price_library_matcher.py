"""
万鼎价格库匹配器 - 将 product_name + spec 匹配到 Material(code) 与价格

Inventory Agent 负责：万鼎匹配 + 库存查询。
数据源：万鼎价格库 管材 sheet，C 列 Describrition（完整中文名）。
方式：CONTAINS + 中文分词 + 向量 fallback。
输出：code、matched_name、unit_price。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, List, Optional

from inventory.config import config

logger = logging.getLogger(__name__)

CACHE_VECTORS_SUFFIX = "_wand_embeddings.npy"

COL_MATERIAL = 1
COL_DESCRIBRITION = 2
PRICE_COLS = {"A": 8, "B": 10, "C": 12, "D": 14}


def _val(v: Any) -> str:
    if v is None:
        return ""
    if hasattr(v, "__class__") and "Formula" in type(v).__name__:
        return ""
    return str(v).strip()


# 规格数字与描述中常见表示的映射（管径，inch 与 mm 对应：1/2"≈15 3/4"≈20 1"≈25）
_SPEC_NUM_PATTERNS = {
    "15": ("dn15", "dn 15", "1/2\"", "(1/2)", "Φ15", "φ15"),
    "20": ("dn20", "dn 20", "3/4\"", "(3/4)", "Φ20", "φ20"),
    "25": ("dn25", "dn 25", "1\"", "(1)", "Φ25", "φ25"),  # 1"≈25mm
    "32": ("dn32", "dn 32", "Φ32", "φ32"),
    "40": ("dn40", "dn 40", "1.1/4\"", "Φ40", "φ40"),
    "50": ("dn50", "dn 50", "2\"", "(2)", "Φ50", "φ50"),
}

# 产品名同义词：询价词 -> 库中词，用于分词时扩展匹配（三角阀→角阀 不添加，业务期望无货）
_SYNONYMS = {"半弯": "弯头"}


def _parse_specs(phrase: str) -> list[str]:
    """从 phrase 解析规格，用于过滤候选。返回如 ['25','dn25'] 或 ['50cm']。"""
    specs: list[str] = []
    phrase = (phrase or "").strip()
    # dn25, dn20, dn50 等
    for m in re.finditer(r"dn\s*(\d+)", phrase, re.I):
        specs.append(f"dn{m.group(1)}")
        specs.append(m.group(1))
    # 50cm, 100cm 等
    for m in re.finditer(r"(\d+)\s*cm", phrase, re.I):
        specs.append(f"{m.group(1)}cm")
        specs.append(m.group(1))
    # 词首数字：25三通、20管卡 → 25、20 作为管径暗示
    lead = re.match(r"^(\d+)[\u4e00-\u9fff]", phrase)
    if lead:
        n = lead.group(1)
        if n not in specs:
            specs.append(n)
    return list(dict.fromkeys(specs))


def _desc_matches_specs(desc: str, specs: list[str]) -> bool:
    """描述是否满足至少一个规格。specs 为空则视为匹配。"""
    if not specs:
        return True
    desc_lower = desc.lower()
    for s in specs:
        s_lower = s.lower()
        if s_lower in desc_lower:
            return True
        # 数字规格的等价形式
        if s in _SPEC_NUM_PATTERNS:
            for p in _SPEC_NUM_PATTERNS[s]:
                if p.lower() in desc_lower:
                    return True
        if s.endswith("cm") and s[:-2].isdigit():
            # 50cm → desc 需含 50 和 cm
            n = s[:-2]
            if n in desc_lower and "cm" in desc_lower:
                return True
    return False


class PriceLibraryMatcher:
    """
    万鼎价格库匹配：phrase → (code, matched_name, unit_price)
    CONTAINS + 中文分词 fallback + 向量 fallback。
    """

    def __init__(
        self,
        price_library_path: str | Path | None = None,
        sheet_name: str = "管材",
        customer_level: str = "B",
    ):
        self.sheet_name = sheet_name
        self.customer_level = (customer_level or "B").upper()
        if self.customer_level not in PRICE_COLS:
            self.customer_level = "B"
        self._price_col = PRICE_COLS[self.customer_level]
        self._rows: list[list] = []
        self._row_data: list[tuple[str, str, float]] = []  # (code, desc, unit_price) per data row
        self._vectors = None
        self._vector_ready = False
        self._loaded = False
        self._path: Optional[Path] = None
        if price_library_path:
            self.load(price_library_path)

    def _resolve_path(self, path: str | Path) -> Path:
        p = Path(path)
        if p.is_absolute() and p.exists():
            return p
        root = Path(__file__).resolve().parent.parent.parent
        return root / p if not p.is_absolute() else p

    def load(self, path: str | Path) -> bool:
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl 未安装，PriceLibraryMatcher 不可用")
            return False
        p = self._resolve_path(path)
        if not p.exists():
            logger.warning(f"万鼎价格库不存在: {p}")
            return False
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else (wb.active or wb[wb.sheetnames[0]])
            self._rows = []
            for row in ws.iter_rows(max_col=16):
                cells = [None] * 16
                for col_idx, c in enumerate(row):
                    if col_idx < 16:
                        cells[col_idx] = getattr(c, "value", None)
                self._rows.append(cells)
            wb.close()
            self._path = p
            self._loaded = True
            self._build_row_data()
            if not self._try_load_vectors_from_cache() and getattr(config, "ENABLE_WANDING_VECTOR", True):
                self._build_vectors()  # 首次加载预构建并缓存，下次可直接读取
            logger.info(f"PriceLibraryMatcher 已加载: {p.name}, {len(self._rows)} 行")
            return True
        except Exception as e:
            logger.warning(f"加载万鼎价格库失败: {e}")
            return False

    def _build_row_data(self) -> None:
        """从 _rows 提取 (code, desc, unit_price) 供向量检索用。"""
        self._row_data = []
        for cells in self._rows[1:]:
            if len(cells) <= COL_DESCRIBRITION:
                self._row_data.append(("", "", 0.0))
                continue
            desc = _val(cells[COL_DESCRIBRITION])
            material = _val(cells[COL_MATERIAL])
            unit_price = 0.0
            if len(cells) > self._price_col and cells[self._price_col] is not None:
                try:
                    unit_price = float(cells[self._price_col])
                except (ValueError, TypeError):
                    pass
            self._row_data.append((material, desc, unit_price))

    def _vector_cache_path(self, table_path: Path) -> Path:
        return table_path.parent / (table_path.stem + CACHE_VECTORS_SUFFIX)

    def _try_load_vectors_from_cache(self) -> bool:
        """从 .npy 缓存加载向量，行数一致则用。"""
        if not self._path or not self._row_data:
            return False
        vec_path = self._vector_cache_path(self._path)
        if not vec_path.is_file():
            return False
        try:
            import numpy as np
            arr = np.load(vec_path)
            if len(arr) != len(self._row_data):
                logger.warning(f"万鼎向量缓存行数 {len(arr)} 与表 {len(self._row_data)} 不一致，将重新预计算")
                return False
            self._vectors = arr
            self._vector_ready = True
            logger.info(f"万鼎向量已从缓存加载: {vec_path}, {len(self._vectors)} 条")
            return True
        except Exception as e:
            logger.warning(f"加载万鼎向量缓存失败: {e}，将重新预计算")
            return False

    def _build_vectors(self) -> None:
        """对 C 列 Describrition 做 embedding，并落盘缓存。行与 _row_data 一一对应。"""
        if not self._row_data or not self._path:
            return
        if not getattr(config, "ENABLE_WANDING_VECTOR", True):
            return
        try:
            import numpy as np
            from openai import OpenAI
            texts = [rd[1] or " " for rd in self._row_data]  # 保证与 _row_data 对齐
            if not any(t.strip() for t in texts):
                return
            timeout = getattr(config, "EMBEDDING_TIMEOUT", 15)
            client = OpenAI(timeout=timeout)
            model = config.OPENAI_EMBEDDING_MODEL
            batch_size = 100
            all_embeds = []
            for i in range(0, len(texts), batch_size):
                chunk = texts[i : i + batch_size]
                r = client.embeddings.create(model=model, input=chunk)
                all_embeds.extend([e.embedding for e in r.data])
            self._vectors = np.array(all_embeds, dtype="float32")
            self._vector_ready = True
            logger.info(f"万鼎向量已预计算: {len(self._vectors)} 条, model={model}")
            self._save_vectors_to_cache()
        except Exception as e:
            logger.warning(f"万鼎向量预计算失败: {e}，将仅使用 CONTAINS + 中文分词")

    def _save_vectors_to_cache(self) -> None:
        """将 _vectors 写入 .npy 缓存。"""
        if self._vectors is None or not self._vector_ready or not self._path:
            return
        vec_path = self._vector_cache_path(self._path)
        try:
            import numpy as np
            np.save(vec_path, self._vectors)
            logger.info(f"万鼎向量已写入缓存: {vec_path}")
        except Exception as e:
            logger.warning(f"写入万鼎向量缓存失败: {e}")

    def _ensure_vectors(self) -> bool:
        """确保向量可用：已有则返回 True；否则尝试加载或构建。"""
        if self._vector_ready and self._vectors is not None:
            return True
        if self._try_load_vectors_from_cache():
            return True
        self._build_vectors()
        return self._vector_ready

    def _match_vector(self, phrase: str) -> Optional[dict[str, Any]]:
        """向量相似度 top_k 检索，返回最佳匹配（无规格过滤）。"""
        return self._match_vector_with_specs(phrase, [])

    def _match_vector_with_specs(self, phrase: str, specs: list[str]) -> Optional[dict[str, Any]]:
        """向量 top_k 检索，有 specs 时优先返回满足规格的，否则返回相似度最高。"""
        if not phrase or not getattr(config, "ENABLE_WANDING_VECTOR", True):
            return None
        if not self._ensure_vectors() or self._vectors is None:
            return None
        try:
            import numpy as np
            from openai import OpenAI
            timeout = getattr(config, "EMBEDDING_TIMEOUT", 15)
            client = OpenAI(timeout=timeout)
            model = config.OPENAI_EMBEDDING_MODEL
            r = client.embeddings.create(model=model, input=[phrase.strip()])
            q = np.array(r.data[0].embedding, dtype="float32")
            top_k = getattr(config, "WANDING_VECTOR_TOP_K", 3)
            # 有规格时扩大候选池以便过滤
            candidate_k = max(top_k, 20) if specs else top_k
            valid_mask = np.array([bool(rd[1]) for rd in self._row_data], dtype=bool)
            if not np.any(valid_mask):
                return None
            sim = np.dot(self._vectors, q)
            sim[~valid_mask] = -1e9
            indices = np.argsort(-sim)[:candidate_k]
            candidates = []
            for idx in indices:
                i = int(idx)
                code, desc, unit_price = self._row_data[i]
                candidates.append({
                    "code": str(code).strip(),
                    "price_row": i + 2,
                    "matched_name": desc,
                    "unit_price": unit_price,
                })
            return self._pick_by_specs(candidates, specs)
        except Exception as e:
            logger.warning(f"万鼎向量检索失败: {e}")
            return None

    def match_vector_candidates(
        self,
        phrase: str,
        min_score: Optional[float] = None,
        max_candidates: Optional[int] = None,
    ) -> List[dict[str, Any]]:
        """
        向量粗筛：用 keywords 做 embedding 相似度检索，返回 score >= min_score 的候选列表。
        向量最泛化，适合作为粗筛阶段产出候选池，供后续规格过滤、业务规则、LLM 选择。
        """
        if not self.is_available() or not (phrase or "").strip():
            return []
        if not getattr(config, "ENABLE_WANDING_VECTOR", True):
            return []
        if not self._ensure_vectors() or self._vectors is None:
            return []

        min_score = min_score if min_score is not None else getattr(config, "WANDING_VECTOR_MIN_SCORE", 0.65)
        max_candidates = max_candidates if max_candidates is not None else getattr(config, "WANDING_VECTOR_COARSE_MAX", 20)

        try:
            import numpy as np
            from openai import OpenAI

            timeout = getattr(config, "EMBEDDING_TIMEOUT", 15)
            client = OpenAI(timeout=timeout)
            model = getattr(config, "OPENAI_EMBEDDING_MODEL", "text-embedding-3-large")
            r = client.embeddings.create(model=model, input=[phrase.strip()])
            q = np.array(r.data[0].embedding, dtype="float32")

            valid_mask = np.array([bool(rd[1]) for rd in self._row_data], dtype=bool)
            if not np.any(valid_mask):
                return []

            sim = np.dot(self._vectors, q)
            sim[~valid_mask] = -1e9

            out: List[dict[str, Any]] = []
            indices = np.argsort(-sim)
            for idx in indices:
                s = float(sim[idx])
                if s < min_score:
                    break
                i = int(idx)
                code, desc, unit_price = self._row_data[i]
                cells = self._rows[i + 1] if i + 1 < len(self._rows) else []
                row_idx = i + 2
                up = 0.0
                if len(cells) > self._price_col and cells[self._price_col] is not None:
                    try:
                        up = float(cells[self._price_col])
                    except (ValueError, TypeError):
                        pass
                out.append({
                    "code": str(code).strip(),
                    "matched_name": desc or "",
                    "unit_price": up,
                    "price_row": row_idx,
                    "source": "vector",
                    "score": round(s, 4),
                })
                if len(out) >= max_candidates:
                    break
            return out
        except Exception as e:
            logger.warning("万鼎向量粗筛失败: %s", e)
            return []

    def is_available(self) -> bool:
        return self._loaded and len(self._rows) > 1

    def _row_to_result(self, row_idx: int, cells: list, material: str, desc: str) -> dict[str, Any]:
        unit_price = 0.0
        if len(cells) > self._price_col and cells[self._price_col] is not None:
            try:
                unit_price = float(cells[self._price_col])
            except (ValueError, TypeError):
                pass
        return {"code": str(material).strip(), "price_row": row_idx, "matched_name": desc, "unit_price": unit_price}

    def _pick_by_specs(self, candidates: list[dict], specs: list[str]) -> Optional[dict[str, Any]]:
        """从候选中按规格过滤，有规格时优先返回满足的，否则返回第一个。"""
        if not candidates:
            return None
        if not specs:
            return candidates[0]
        for c in candidates:
            if _desc_matches_specs(c.get("matched_name", ""), specs):
                return c
        return candidates[0]  # 规格过滤无命中时退回首个候选

    def match(self, phrase: str) -> Optional[dict[str, Any]]:
        if not self.is_available() or not (phrase or "").strip():
            return None
        phrase = phrase.strip()
        phrase_lower = phrase.lower()
        specs = _parse_specs(phrase)

        # 1. CONTAINS
        contains_cands: list[dict] = []
        for row_idx, cells in enumerate(self._rows[1:], start=2):
            if len(cells) <= COL_DESCRIBRITION:
                continue
            desc = _val(cells[COL_DESCRIBRITION]) if COL_DESCRIBRITION < len(cells) else ""
            material = _val(cells[COL_MATERIAL]) if COL_MATERIAL < len(cells) else ""
            if not desc:
                continue
            if phrase_lower in desc.lower():
                contains_cands.append(self._row_to_result(row_idx, cells, material, desc))
        if contains_cands:
            return self._pick_by_specs(contains_cands, specs)

        # 2. 中文分词（含同义词扩展）
        if re.search(r"[\u4e00-\u9fff]", phrase):
            chinese_chars = re.findall(r"[\u4e00-\u9fff]+", phrase)
            if chinese_chars:
                longest = max(chinese_chars, key=len)
                if len(longest) >= 2:
                    keywords = [longest[:2], longest[-2:]] if len(longest) >= 4 else [longest]
                    keywords = list(dict.fromkeys(keywords))
                    # 同义词：用库中词替换以匹配（如 半弯→弯头）
                    if longest in _SYNONYMS:
                        keywords = [_SYNONYMS[longest]]
                    seg_cands: list[dict] = []
                    for row_idx, cells in enumerate(self._rows[1:], start=2):
                        if len(cells) <= COL_DESCRIBRITION:
                            continue
                        desc = _val(cells[COL_DESCRIBRITION]) if COL_DESCRIBRITION < len(cells) else ""
                        material = _val(cells[COL_MATERIAL]) if COL_MATERIAL < len(cells) else ""
                        if not desc:
                            continue
                        desc_lower = desc.lower()
                        if all(kw.lower() in desc_lower for kw in keywords):
                            seg_cands.append(self._row_to_result(row_idx, cells, material, desc))
                    if seg_cands:
                        return self._pick_by_specs(seg_cands, specs)

        # 3. 向量：扩大候选池并做规格过滤
        if getattr(config, "ENABLE_WANDING_VECTOR", True):
            vec_res = self._match_vector_with_specs(phrase, specs)
            if vec_res:
                return vec_res
        return None

    def match_by_code(self, code: str) -> Optional[dict[str, Any]]:
        """按 Material(code) 精确匹配，返回 matched_name、unit_price。"""
        if not self.is_available() or not (code or "").strip():
            return None
        code_str = str(code).strip()
        for row_idx, cells in enumerate(self._rows[1:], start=2):
            if len(cells) <= COL_MATERIAL:
                continue
            material = _val(cells[COL_MATERIAL])
            if material and str(material).strip() == code_str:
                desc = _val(cells[COL_DESCRIBRITION]) if len(cells) > COL_DESCRIBRITION else ""
                unit_price = 0.0
                if len(cells) > self._price_col and cells[self._price_col] is not None:
                    try:
                        unit_price = float(cells[self._price_col])
                    except (ValueError, TypeError):
                        pass
                return {"code": code_str, "price_row": row_idx, "matched_name": desc, "unit_price": unit_price}
        return None
