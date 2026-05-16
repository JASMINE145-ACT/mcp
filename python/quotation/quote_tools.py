# 报价 Agent 工具：从原始报价单 Excel 提取「第2行到 Total Excluding PPN不含税总价 所在行之前」的报价数据，供 LLM 使用

from __future__ import annotations

import json
import logging
import os
import re
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any, List, Tuple

logger = logging.getLogger(__name__)

# 表头关键词：交货日期列、报价日期标签（用于定位填写格）
DELIVERY_DATE_COL_KEYWORDS = [
    "交货日期",
    "Tanggal Pengiriman",
    "Delivery Date",
    "Tanggal Pengiriman Barang",
]
QUOTATION_DATE_LABEL_KEYWORDS = [
    "报价日期",
    "Tanggal Penawaran",
    "Quotation Date",
]

# 凌威模板：footer 中「报价日期」标签常在 J:K 合并格，数值应在 L:R 合并格（旧逻辑写 c+1 会落在 J:K 内）。
QUOTE_DATE_VALUE_MIN_COL = 12  # L
QUOTE_DATE_VALUE_MAX_COL = 18  # R


def _copy_cell_style(source_cell, dest_cell) -> None:
    """
    将 source_cell 的格式（边框、底纹、字体、对齐、数字格式）复制到 dest_cell，
    避免填表时新建单元格使用默认样式导致虚线边框或异常底纹（如绿色块）。
    """
    try:
        if getattr(source_cell, "has_style", False) and getattr(source_cell, "_style", None) is not None:
            dest_cell._style = copy(source_cell._style)
            return
    except (TypeError, AttributeError):
        pass
    try:
        if getattr(source_cell, "font", None) is not None:
            dest_cell.font = copy(source_cell.font)
        if getattr(source_cell, "border", None) is not None:
            dest_cell.border = copy(source_cell.border)
        if getattr(source_cell, "fill", None) is not None:
            dest_cell.fill = copy(source_cell.fill)
        if getattr(source_cell, "alignment", None) is not None:
            dest_cell.alignment = copy(source_cell.alignment)
        if getattr(source_cell, "number_format", None) is not None:
            dest_cell.number_format = source_cell.number_format
    except (TypeError, AttributeError):
        logger.debug("复制单元格样式时部分属性失败", exc_info=True)


def _normalize_sheet_view(ws) -> None:
    """
    Normalize worksheet view for exported files.
    Some templates are saved in page-break preview mode, which users perceive
    as dashed lines and large tinted areas in Excel/WPS.
    Also clears conditional formatting and removes green-ish fills from the
    entire worksheet to eliminate template artifacts.
    """
    try:
        if hasattr(ws, "views") and hasattr(ws.views, "sheetView") and ws.views.sheetView:
            for sv in ws.views.sheetView:
                sv.view = "normal"
        if getattr(ws, "sheet_view", None) is not None:
            ws.sheet_view.view = "normal"
    except Exception:
        logger.debug("normalize sheet view failed", exc_info=True)

    try:
        ws.page_setup.horizontalCentered = False
        ws.page_setup.verticalCentered = False
        ws.sheet_properties.pageSetUpPr = None
    except Exception:
        logger.debug("clear page setup properties failed", exc_info=True)

    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        logger.debug("clear conditional formatting failed", exc_info=True)

    try:
        from openpyxl.styles import PatternFill
        no_fill = PatternFill(fill_type=None)
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                fill = getattr(cell, "fill", None)
                if fill is None:
                    continue
                fg = getattr(fill.fgColor, "rgb", None) if fill.fgColor else None
                bg = getattr(fill.bgColor, "rgb", None) if fill.bgColor else None
                if fg and isinstance(fg, str) and _is_green_like(fg):
                    cell.fill = no_fill
                elif bg and isinstance(bg, str) and _is_green_like(bg):
                    cell.fill = no_fill
    except Exception:
        logger.debug("clear green fills failed", exc_info=True)


def _is_green_like(rgb_hex: str) -> bool:
    """Return True if the RGB hex string looks green-ish (high G, low R/B)."""
    hex_str = rgb_hex.lstrip("#")
    if len(hex_str) == 8:
        hex_str = hex_str[2:]
    if len(hex_str) != 6:
        return False
    try:
        r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
    except ValueError:
        return False
    return g > 150 and g > r * 1.3 and g > b * 1.3

# 边界行标识（用户指定：报价数据列到该行为止）
TOTAL_ROW_MARKER = "Total Excluding PPN不含税总价"

# 询价列表头关键词（用于自动识别列）
NAME_COL_KEYWORDS = ["询价货物名称", "Nama Permintaan Barang", "nama permintaan"]
SPEC_COL_KEYWORDS = ["询价规格型号", "Spesifikasi dan Model Permintaan Barang", "Spesifikasi"]
QTY_COL_KEYWORDS = ["Jumlah", "数量", "jumlah", "Quantity"]


def _cell_value(cell) -> str:
    v = getattr(cell, "value", None)
    if v is None:
        return ""
    return str(v).strip()


def _to_int_or_none(value: Any) -> int | None:
    """Best-effort int conversion; return None on invalid input."""
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _to_float_or_none(value: Any) -> float | None:
    """Best-effort float conversion; return None on invalid input."""
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _set_cell_value_merged_safe(ws, row: int, col: int, value: Any) -> None:
    """
    在处理合并单元格时安全地写入值：
    - 若目标单元格为普通单元格，直接写入；
    - 若为合并区域中的「非左上角」MergedCell，则改为写入该合并区域的左上角单元格，
      避免 openpyxl 抛出 "'MergedCell' object attribute 'value' is read-only"。
    """
    try:
        from openpyxl.cell.cell import MergedCell  # type: ignore
    except Exception:  # pragma: no cover - 仅防御性兜底
        ws.cell(row=row, column=col, value=value)
        return

    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    # 定位包含该 MergedCell 的合并区域，并将值写入左上角单元格
    try:
        for merged_range in getattr(ws, "merged_cells", []).ranges:  # type: ignore[attr-defined]
            if (merged_range.min_row <= row <= merged_range.max_row) and (
                merged_range.min_col <= col <= merged_range.max_col
            ):
                master = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                master.value = value
                return
    except Exception:
        # 回退到直接写入目标单元格，若再次触发异常则交由调用方处理
        pass
    cell.value = value


def extract_quotation_data(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    从原始报价单 Excel 提取报价数据：从第 2 行起到「Total Excluding PPN不含税总价」所在行的上一行止。

    - 第 1 行视为表头。
    - 数据区：第 2 行 ～ 第一个包含 TOTAL_ROW_MARKER 的行的上一行。
    - 支持 .xlsx / .xlsm；.xls 需调用方先转为 xlsx 或在此用 pandas 扩展。

    Returns:
        {"success": bool, "result": str, "error": str | None, "rows_count": int}
        result 为 Markdown 表格或 JSON 文本，便于 LLM 阅读。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "result": "", "error": "仅支持 .xlsx / .xlsm；.xls 请先另存为 xlsx", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "result": "", "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        # 先收集所有行（read_only 下 iter_rows 只能遍历一次，先找边界行）
        all_rows: List[List[str]] = []
        total_row_1based: int | None = None

        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            cells = [_cell_value(c) for c in row]
            all_rows.append(cells)

            # 任意单元格包含边界标识即视为「总价行」，数据区在其上一行结束
            for c in cells:
                if TOTAL_ROW_MARKER in (c or ""):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            logger.debug("关闭 workbook 失败", exc_info=True)
        return {"success": False, "result": "", "error": str(e), "rows_count": 0}

    if not all_rows:
        return {"success": True, "result": "表为空，无报价数据。", "error": None, "rows_count": 0}

    # 表头：第 1 行（索引 0）
    header = all_rows[0]
    # 数据行：第 2 行到「Total Excluding PPN」的上一行（不包含总价行）
    if total_row_1based is not None and total_row_1based >= 2:
        data_rows = all_rows[1 : total_row_1based - 1]
    else:
        data_rows = all_rows[1:]

    if not data_rows:
        return {"success": True, "result": "未找到数据行（或仅含总价行）。", "error": None, "rows_count": 0}

    # 输出为 Markdown 表格，便于 LLM 理解
    def escape_md(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ")

    col_count = max(len(header), max(len(r) for r in data_rows), 1)
    header_padded = header + [""] * (col_count - len(header))
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    lines = ["| " + " | ".join(escape_md(h) for h in header_padded) + " |", sep]
    for r in data_rows:
        r_padded = r + [""] * (col_count - len(r))
        lines.append("| " + " | ".join(escape_md(c) for c in r_padded) + " |")

    result_text = (
        f"报价数据（第2行至「Total Excluding PPN不含税总价」上一行，共 {len(data_rows)} 行；"
        "其中 Qty=询价数量/采购数量，不是库存）：\n\n"
        + "\n".join(lines)
        + f"\n\n（共 {len(data_rows)} 行。回复用户时请列出上表**全部**数据行，勿只列部分。）"
    )
    return {
        "success": True,
        "result": result_text,
        "error": None,
        "rows_count": len(data_rows),
    }


def _find_col_by_header(header: List[str], keywords_list: List[str]) -> int:
    """按表头关键词查找列索引，返回 -1 表示未找到。"""
    for i, cell in enumerate(header):
        val = (str(cell or "").strip()).lower()
        for kw in keywords_list:
            if kw.lower() in val:
                return i
    return -1


def _find_delivery_date_column(ws, max_header_rows: int = 4, max_cols: int = 30) -> int:
    """
    在前几行表头中按关键词查找「交货日期」列，返回 1-based 列号，未找到返回 0。
    """
    for row_1based in range(1, max_header_rows + 1):
        row_cells: List[str] = []
        for c in range(1, max_cols + 1):
            try:
                v = ws.cell(row=row_1based, column=c).value
                row_cells.append(str(v).strip() if v is not None else "")
            except Exception:
                break
        col0 = _find_col_by_header(row_cells, DELIVERY_DATE_COL_KEYWORDS)
        if col0 >= 0:
            return col0 + 1
    return 0


def _find_quotation_date_cell(
    ws, total_row_1based: int, search_rows: int = 20, max_cols: int = 30
) -> Tuple[int, int] | None:
    """
    在合计行下方的 footer 区域查找「报价日期」标签所在行，返回应填写日期的单元格 (row, col) 1-based。

    - 凌威报价单：标签在 J:K 合并格时，旧约定「标签右侧一格」仍落在 J:K 内，错误。
      若同一行存在列号在 L–R（12–18）内的合并区域，则日期写入该区域左上角。
    - 其他模板：无 L–R 合并时，回退为标签所在单元格的右侧一列 (c+1)。
    """
    label_row: int | None = None
    label_col: int | None = None
    for r in range(total_row_1based + 4, total_row_1based + search_rows + 1):
        for c in range(1, max_cols + 1):
            try:
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                s = str(val).strip()
                if any(kw in s for kw in QUOTATION_DATE_LABEL_KEYWORDS):
                    label_row, label_col = r, c
                    break
            except Exception:
                continue
        if label_row is not None:
            break
    if label_row is None or label_col is None:
        return None

    try:
        candidates: list = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= label_row <= merged_range.max_row:
                if (
                    merged_range.min_col >= QUOTE_DATE_VALUE_MIN_COL
                    and merged_range.min_col <= QUOTE_DATE_VALUE_MAX_COL
                ):
                    candidates.append(merged_range)
        single_row = [m for m in candidates if m.min_row == m.max_row == label_row]
        if single_row:
            prefer_l = [m for m in single_row if m.min_col == QUOTE_DATE_VALUE_MIN_COL]
            chosen = prefer_l[0] if prefer_l else min(single_row, key=lambda m: m.min_col)
            return (label_row, chosen.min_col)
        if candidates:
            on_row = [m for m in candidates if m.min_row <= label_row <= m.max_row]
            if on_row:
                chosen = min(on_row, key=lambda m: (m.min_row, m.min_col))
                return (label_row, chosen.min_col)
    except Exception:
        logger.debug("_find_quotation_date_cell merge scan failed", exc_info=True)

    return (label_row, label_col + 1)


def _extract_inquiry_items_smart_fallback(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 200,
) -> dict[str, Any]:
    """
    普适解析 fallback：不依赖 TOTAL_ROW_MARKER，读取工作表前 max_rows 行，
    在前 3 行中按关键词识别名称/规格/数量列，构建 items。供 extract_inquiry_items 在主逻辑失败或无数据时调用。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "items": [], "error": "请安装 openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "items": [], "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "items": [], "error": "仅支持 .xlsx / .xlsm", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "items": [], "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        rows: List[List[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(c).strip() if c is not None else "" for c in (row or [])])
        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            logger.debug("关闭 workbook 失败", exc_info=True)
        return {"success": False, "items": [], "error": str(e), "rows_count": 0}

    if not rows or len(rows) < 2:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    name_col = spec_col = qty_col = -1
    header_row_idx = 0
    for idx, header_row in enumerate(rows[:3]):
        nc = _find_col_by_header(header_row, NAME_COL_KEYWORDS)
        sc = _find_col_by_header(header_row, SPEC_COL_KEYWORDS)
        qc = _find_col_by_header(header_row, QTY_COL_KEYWORDS)
        if nc >= 0:
            name_col, spec_col = nc, sc
            if qc >= 0:
                qty_col = qc
            header_row_idx = idx
            break

    if name_col < 0:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    data_rows = rows[header_row_idx + 1:]
    items: List[dict] = []
    for i, row_cells in enumerate(data_rows):
        row_num = header_row_idx + 2 + i
        product_name = (row_cells[name_col] if name_col < len(row_cells) else "").strip()
        specification = (row_cells[spec_col] if spec_col >= 0 and spec_col < len(row_cells) else "").strip()
        keywords = f"{product_name} {specification}".strip() if specification else product_name
        if not keywords:
            continue
        qty_val = 0
        if qty_col >= 0 and qty_col < len(row_cells):
            try:
                v = row_cells[qty_col]
                if v is not None and str(v).strip():
                    qty_val = int(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                pass
        items.append({
            "row": row_num,
            "product_name": product_name,
            "specification": specification,
            "keywords": keywords,
            "qty": qty_val,
        })

    return {"success": True, "items": items, "error": None, "rows_count": len(items)}


def extract_inquiry_items(
    file_path: str,
    sheet_name: str | None = None,
    col_mapping: dict | None = None,
) -> dict[str, Any]:
    """
    提取「询价货物名称」「询价规格型号」两列，输出供库存查询 Agent 用的列表。

    - 复用 data 区域识别逻辑（第2行～Total Excluding PPN 上一行）
    - 列识别：按表头匹配，或通过 col_mapping 指定 {name_col: int, spec_col: int}
    - keywords = product_name + " " + specification（空规格则仅名称）

    Returns:
        {"success": bool, "items": [...], "error": str | None, "rows_count": int}
        items: [{"row": 1, "product_name": "", "specification": "", "keywords": ""}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "items": [], "error": "请安装 openpyxl: pip install openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "items": [], "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "items": [], "error": "仅支持 .xlsx / .xlsm", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "items": [], "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        all_rows: List[List[str]] = []
        total_row_1based: int | None = None

        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            cells = [_cell_value(c) for c in row]
            all_rows.append(cells)
            for c in cells:
                if TOTAL_ROW_MARKER in (c or ""):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            logger.debug("关闭 workbook 失败", exc_info=True)
        return {"success": False, "items": [], "error": str(e), "rows_count": 0}

    if not all_rows:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    # 确定表头行与列索引（部分模板首行为标题，次行为表头）
    if col_mapping:
        name_col = col_mapping.get("name_col", col_mapping.get("product_name_col", -1))
        spec_col = col_mapping.get("spec_col", col_mapping.get("specification_col", -1))
        qty_col = col_mapping.get("qty_col", col_mapping.get("quantity_col", -1))
        header_row_idx = 0
    else:
        name_col = spec_col = qty_col = -1
        header_row_idx = 0
        for idx, header_row in enumerate(all_rows[:3]):
            nc = _find_col_by_header(header_row, NAME_COL_KEYWORDS)
            sc = _find_col_by_header(header_row, SPEC_COL_KEYWORDS)
            qc = _find_col_by_header(header_row, QTY_COL_KEYWORDS)
            if nc >= 0:
                name_col, spec_col = nc, sc
                if qc >= 0:
                    qty_col = qc
                header_row_idx = idx
                break

    if name_col < 0:
        # Fallback：用普适解析（不依赖 Total Excluding PPN 与固定表头）再尝试识别列
        fallback = _extract_inquiry_items_smart_fallback(file_path, sheet_name)
        if fallback.get("items"):
            fallback["_fallback_used"] = True
            fallback["error"] = None
            return fallback
        return {"success": False, "items": [], "error": "未找到询价货物名称列，请检查表头或提供 col_mapping", "rows_count": 0}

    # 数据行从表头下一行起，到 Total Excluding PPN 上一行
    data_start = header_row_idx + 1
    if total_row_1based is not None and total_row_1based >= 2:
        data_end = total_row_1based - 1
    else:
        data_end = len(all_rows)
    data_rows = all_rows[data_start:data_end]

    if not data_rows:
        fallback = _extract_inquiry_items_smart_fallback(file_path, sheet_name)
        if fallback.get("items"):
            fallback["_fallback_used"] = True
            return fallback
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    # spec_col 可为 -1，表示无规格列；qty_col 可为 -1，表示无数量列
    items: List[dict] = []
    for i, row_cells in enumerate(data_rows):
        row_num = data_start + 1 + i  # Excel 行号 1-based
        product_name = (row_cells[name_col] if name_col < len(row_cells) else "").strip()
        specification = (row_cells[spec_col] if spec_col >= 0 and spec_col < len(row_cells) else "").strip()
        keywords = f"{product_name} {specification}".strip() if specification else product_name
        if not keywords:
            continue
        # 需求数量 qty：用于库存比对
        qty_val = 0
        if qty_col >= 0 and qty_col < len(row_cells):
            try:
                v = row_cells[qty_col]
                if v is not None and str(v).strip():
                    qty_val = int(float(str(v).replace(",", "")))
            except (ValueError, TypeError) as e:
                logger.debug("解析数量失败 row=%s: %s", row_num, e)
        items.append({
            "row": row_num,
            "product_name": product_name,
            "specification": specification,
            "keywords": keywords,
            "qty": qty_val,
        })

    return {
        "success": True,
        "items": items,
        "error": None,
        "rows_count": len(items),
    }


# 案例报价单模板：表头第 2 行，数据从第 3 行起，A=序号 B=询价货物名称 C=规格 E=数量
INQUIRY_DATA_START_ROW = 3
INQUIRY_COL_SEQ = 1   # A
INQUIRY_COL_NAME = 2  # B
INQUIRY_COL_SPEC = 3  # C
INQUIRY_COL_QTY = 5   # E


def fill_template_with_inquiry_items(
    template_path: str,
    items: List[dict[str, Any]],
    output_path: str,
    sheet_name: str = "询价单",
    allow_insert_rows: bool = False,
) -> dict[str, Any]:
    """
    用「询价行」列表填写案例报价单模板，生成可被 extract_inquiry_items / Work 流程处理的 Excel。

    - 复制模板到 output_path，不修改原模板。
    - 从第 3 行起写入：A=序号(1-based)，B=product_name，C=specification，E=qty。
    - 默认不插行（allow_insert_rows=False），只在模板可用行内写入，确保文档样式稳定。
    - 如需扩展行数可启用 allow_insert_rows=True（复杂模板可能导致样式/结构异常）。

    Returns:
        {"success": bool, "output_path": str, "filled_count": int, "error": str | None, "capacity": int, "truncated_count": int}
    """
    import shutil
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "output_path": "", "filled_count": 0, "error": "请安装 openpyxl"}

    tpl = Path(template_path)
    if not tpl.is_absolute():
        tpl = Path(os.getcwd()) / tpl
    if not tpl.exists():
        return {"success": False, "output_path": "", "filled_count": 0, "error": f"模板不存在: {tpl}"}
    out_p = Path(output_path)
    if not out_p.is_absolute():
        out_p = Path(os.getcwd()) / out_p
    out_p.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(tpl, out_p)
    except Exception as e:
        return {"success": False, "output_path": "", "filled_count": 0, "error": str(e)}

    items = [x for x in items if isinstance(x, dict) and (x.get("product_name") or x.get("name"))]
    if not items:
        return {
            "success": True,
            "output_path": str(out_p),
            "filled_count": 0,
            "error": None,
            "capacity": 0,
            "truncated_count": 0,
        }

    try:
        wb = openpyxl.load_workbook(out_p)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        total_row_1based = None
        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            for cell in row:
                if TOTAL_ROW_MARKER in _cell_value(cell):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break
        if total_row_1based is None:
            total_row_1based = ws.max_row + 1
        data_start = INQUIRY_DATA_START_ROW
        available = max(0, total_row_1based - data_start)
        truncated_count = 0
        if len(items) > available:
            if allow_insert_rows:
                insert_count = len(items) - available
                ws.insert_rows(total_row_1based, insert_count)
                # 新插入行从上一行复制样式，避免出现默认虚线/绿底
                style_row = total_row_1based - 1
                for new_row in range(total_row_1based, total_row_1based + insert_count):
                    for col in range(1, min(ws.max_column + 1, 20)):
                        _copy_cell_style(ws.cell(row=style_row, column=col), ws.cell(row=new_row, column=col))
            else:
                truncated_count = len(items) - available
                items = items[:available]
        filled = 0
        for i, it in enumerate(items):
            row_num = data_start + i
            name = (it.get("product_name") or it.get("name") or "").strip()
            spec = (it.get("specification") or it.get("spec") or "").strip()
            qty = _to_int_or_none(it.get("qty", 0))
            if qty is None:
                qty = 0
            ws.cell(row=row_num, column=INQUIRY_COL_SEQ, value=i + 1)
            ws.cell(row=row_num, column=INQUIRY_COL_NAME, value=name)
            ws.cell(row=row_num, column=INQUIRY_COL_SPEC, value=spec)
            ws.cell(row=row_num, column=INQUIRY_COL_QTY, value=max(0, qty))
            filled += 1
        _normalize_sheet_view(ws)
        wb.save(out_p)
        return {
            "success": True,
            "output_path": str(out_p),
            "filled_count": filled,
            "error": None,
            "capacity": available,
            "truncated_count": truncated_count,
        }
    except Exception as e:
        return {
            "success": False,
            "output_path": "",
            "filled_count": 0,
            "error": str(e),
            "capacity": 0,
            "truncated_count": 0,
        }


def get_template_inquiry_capacity(
    template_path: str,
    sheet_name: str = "询价单",
) -> dict[str, Any]:
    """
    读取询价模板在不插行模式下的可填写容量（可写行数）。
    capacity = Total 行号 - INQUIRY_DATA_START_ROW
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "capacity": 0, "error": "请安装 openpyxl"}

    path = Path(template_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "capacity": 0, "error": f"模板不存在: {path}"}

    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        total_row_1based = None
        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            for cell in row:
                if TOTAL_ROW_MARKER in _cell_value(cell):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break
        if total_row_1based is None:
            total_row_1based = ws.max_row + 1
        capacity = max(0, total_row_1based - INQUIRY_DATA_START_ROW)
        wb.close()
        return {"success": True, "capacity": capacity, "error": None}
    except Exception as e:
        return {"success": False, "capacity": 0, "error": str(e)}


# ---------------------------------------------------------------------------
# 普适性 Excel 工具（不依赖报价单结构，任意 Excel 可用）
# ---------------------------------------------------------------------------

def parse_excel_smart(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 500,
) -> dict[str, Any]:
    """
    【普适性】智能解析任意 Excel：自动读取指定工作表的所有单元格（或前 max_rows 行），
    零硬编码列/行，适合多表头、合并单元格、不规则布局。返回 Markdown 表格便于 LLM 理解。

    Returns:
        {"success": bool, "result": str, "error": str | None, "sheet_name": str, "rows_read": int}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "sheet_name": "", "rows_read": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "sheet_name": "", "rows_read": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "result": "", "error": "仅支持 .xlsx / .xlsm", "sheet_name": "", "rows_read": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "result": "", "error": f"打开 Excel 失败: {e}", "sheet_name": "", "rows_read": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            used_sheet = sheet_name
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
            used_sheet = ws.title

        rows: List[List[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(c).strip() if c is not None else "" for c in (row or [])])
        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"success": False, "result": "", "error": str(e), "sheet_name": used_sheet if "used_sheet" in dir() else "", "rows_read": 0}

    if not rows:
        return {"success": True, "result": "表为空或无数据。", "error": None, "sheet_name": used_sheet, "rows_read": 0}

    def escape_md(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ")

    col_count = max(len(r) for r in rows)
    lines = []
    for i, r in enumerate(rows):
        r_padded = list(r) + [""] * (col_count - len(r))
        lines.append("| " + " | ".join(escape_md(str(c)) for c in r_padded) + " |")
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    header_block = (
        f"工作表「{used_sheet}」共 {len(rows)} 行（普适解析，未限定列）：\n\n"
        + "| " + " | ".join(escape_md(str(i + 1)) for i in range(col_count)) + " |\n"
        + sep + "\n"
    )
    # 若结果过长，在完整行边界处截断，避免下游在行中间截断导致模型在单元格填「数据被截断」
    _max_result_chars = 40_000
    if len(header_block) + sum(len(ln) + 1 for ln in lines) + 200 > _max_result_chars:
        n_show = 0
        acc = len(header_block) + 200
        for ln in lines:
            if acc + len(ln) + 1 > _max_result_chars:
                break
            acc += len(ln) + 1
            n_show += 1
        lines_show = lines[:n_show] if n_show else lines[:1]
        result_text = (
            header_block + "\n".join(lines_show)
            + f"\n\n（因长度限制仅展示前 {len(lines_show)} 行，共 {len(rows)} 行。回复时按上表逐行照抄，勿在单元格内填「数据被截断」。）"
        )
    else:
        result_text = (
            header_block + "\n".join(lines)
            + f"\n\n（共 {len(rows)} 行。回复时**必须按上表逐行照抄**，不得只列部分、不得将同一行重复多遍凑数、不得自行编造行、勿在单元格内填「数据被截断」。）"
        )
    return {"success": True, "result": result_text, "error": None, "sheet_name": used_sheet, "rows_read": len(rows)}


def _parse_cell_ref(ref: str) -> tuple[int, int] | None:
    """将 A1 形式转为 (row: 1-based, col: 1-based)，失败返回 None。"""
    m = re.match(r"^([A-Za-z]+)(\d+)$", (ref or "").strip())
    if not m:
        return None
    col_str, row_str = m.group(1).upper(), m.group(2)
    col = 0
    for c in col_str:
        col = col * 26 + (ord(c) - ord("A") + 1)
    try:
        row = int(row_str)
        return (row, col) if row >= 1 and col >= 1 else None
    except ValueError:
        return None


def edit_excel(
    file_path: str,
    edits: List[dict[str, Any]],
    sheet_name: str | None = None,
    output_path: str | None = None,
) -> dict[str, Any]:
    """
    【普适性】编辑任意 Excel：按单元格或区域写入。不依赖报价单列结构，任意 .xlsx/.xlsm 可用。

    edits 每项为以下之一：
    - {"cell": "A1", "value": 任意}：单格写入
    - {"range": "A1:B2", "values": [[v1,v2],[v3,v4]]}：区域按行写入

    Returns:
        {"success": bool, "result": str, "error": str | None, "output_path": str}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "output_path": ""}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "output_path": ""}
    out_p = Path(output_path) if output_path else path
    if not out_p.is_absolute():
        out_p = Path(os.getcwd()) / out_p

    if not edits or not isinstance(edits, list):
        return {"success": False, "result": "", "error": "请提供 edits 数组（每项含 cell+value 或 range+values）", "output_path": ""}

    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        applied = 0
        for op in edits:
            if op.get("cell") is not None and "value" in op:
                cell_ref = str(op["cell"]).strip()
                parsed = _parse_cell_ref(cell_ref)
                if not parsed:
                    continue
                row, col = parsed
                ws.cell(row=row, column=col, value=op["value"])
                applied += 1
            elif op.get("range") is not None and op.get("values") is not None:
                range_ref = str(op["range"]).strip()
                parts = range_ref.split(":")
                if len(parts) != 2:
                    continue
                start = _parse_cell_ref(parts[0])
                end = _parse_cell_ref(parts[1])
                if not start or not end:
                    continue
                row_s, col_s = start
                row_e, col_e = end
                vals = op["values"]
                if not isinstance(vals, list):
                    continue
                for ri, row_vals in enumerate(vals):
                    if not isinstance(row_vals, list):
                        continue
                    for ci, v in enumerate(row_vals):
                        r, c = row_s + ri, col_s + ci
                        if r > row_e or c > col_e:
                            break
                        ws.cell(row=r, column=c, value=v)
                applied += 1
        wb.save(out_p)
        return {"success": True, "result": json.dumps({"applied_edits": applied, "output_path": str(out_p)}, ensure_ascii=False), "error": None, "output_path": str(out_p)}
    except Exception as e:
        return {"success": False, "result": "", "error": str(e), "output_path": ""}


def get_quote_tools_openai_format() -> list[dict]:
    """OpenAI function calling 格式：报价 Agent 工具。仅暴露 parse_excel_smart 做 Excel 解析，不再暴露 extract_quotation_data。"""
    return [
        {
            "type": "function",
            "function": {
                "name": "fill_quotation_sheet",
                "description": "【报价单导向】将数据写入报价单 Excel 指定行。fill_items 每项含 row、code、quote_name、unit_price、qty、specification。写入列：G=产品编号, H=报价名称, J=规格, L=数量, N=单价, O=总价；并按表头自动填写「交货日期」「报价日期」（不传则用当天 YYYY/MM/DD）。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "报价单 Excel 路径"},
                        "fill_items": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "row": {"type": "integer", "description": "Excel 行号 1-based"},
                                    "code": {"type": "string"},
                                    "quote_name": {"type": "string"},
                                    "unit_price": {"type": "number"},
                                    "qty": {"type": "integer"},
                                    "specification": {"type": "string"},
                                },
                                "required": ["row"],
                            },
                            "description": "要回填的项列表",
                        },
                        "output_path": {"type": "string", "description": "可选，输出路径，默认覆盖原文件"},
                        "sheet_name": {"type": "string", "description": "工作表名，不传用第一个"},
                        "quotation_date": {"type": "string", "description": "报价日期，如 2026/03/11，不传用当天"},
                        "delivery_date": {"type": "string", "description": "交货日期（每行同一值），如 2026/03/20，不传用当天"},
                    },
                    "required": ["file_path", "fill_items"],
                },
                "x_tool_meta": {"access_mode": "write", "risk_level": "medium", "deferred": True},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "parse_excel_smart",
                "description": "【普适性，推荐】解析任意 Excel：按行读取全表（默认最多 500 行），返回完整 Markdown 表。提取/查看报价单或 Excel 数据时优先使用此工具，可拿到全表行数，不受「Total」行位置影响。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "Excel 文件完整路径（可从 context.file_path 获取）"},
                        "sheet_name": {"type": "string", "description": "工作表名称，不传则使用第一个工作表"},
                        "max_rows": {"type": "integer", "description": "最多读取行数，默认 500", "default": 500},
                    },
                    "required": ["file_path"],
                },
                "x_tool_meta": {"access_mode": "read", "risk_level": "low", "deferred": True},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "edit_excel",
                "description": "【普适性】编辑任意 Excel：按单元格或区域写入，不依赖报价单列结构。edits 每项：{\"cell\": \"A1\", \"value\": 任意} 单格写入，或 {\"range\": \"A1:B2\", \"values\": [[v1,v2],[v3,v4]]} 区域按行写入。可写多格后保存到 output_path 或覆盖原文件。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "要编辑的 Excel 路径"},
                        "edits": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "cell": {"type": "string", "description": "单格引用，如 A1"},
                                    "value": {"description": "写入的值（字符串或数字）"},
                                    "range": {"type": "string", "description": "区域引用，如 A1:B2"},
                                    "values": {
                                        "type": "array",
                                        "items": {
                                            "type": "array",
                                            "items": {
                                                "oneOf": [
                                                    {"type": "string"},
                                                    {"type": "number"},
                                                    {"type": "boolean"},
                                                    {"type": "null"}
                                                ]
                                            }
                                        },
                                        "description": "二维数组，按行写入，每个单元格为字符串/数字/布尔或 null",
                                    },
                                },
                            },
                            "description": "编辑操作列表：单格用 cell+value，区域用 range+values",
                        },
                        "sheet_name": {"type": "string", "description": "工作表名，不传用第一个"},
                        "output_path": {"type": "string", "description": "保存路径，不传则覆盖原文件"},
                    },
                    "required": ["file_path", "edits"],
                },
                "x_tool_meta": {"access_mode": "write", "risk_level": "medium", "deferred": True},
            },
        },
    ]


# 凌威报价单回填列（1-based）
COL_PRODUCT_NO = 7   # G 产品编号
COL_QUOTE_NAME = 8   # H 报价名称
COL_QUOTE_SPEC = 10  # J 报价产品规格
COL_QTY_OUT = 12     # L 数量
COL_UNIT_PRICE = 14  # N 单价
COL_TOTAL = 15       # O 总价

# 4 个价格计算行（在 Total Excluding PPN 所在行及其后 3 行），金额写入列与 COL_TOTAL 一致
TOTALS_VALUE_COL = COL_TOTAL  # O 列

def fill_quotation(
    file_path: str,
    fill_items: list[dict[str, Any]],
    sheet_name: str | None = None,
    output_path: str | None = None,
    freight: float = 0.0,
    quotation_date: str | None = None,
    delivery_date: str | None = None,
) -> dict[str, Any]:
    """
    将匹配到的产品信息回填到报价单 Excel。
    每行保证填写的列：G=产品编号, H=报价名称, J=报价产品规格(无 specification 时用 quote_name), L=数量, N=单价, O=总价；
    未匹配项写「无货」；写完后更新底部 4 个价格计算行；并按表头填写「交货日期」「报价日期」。

    推荐入参：fill_items 来自「规范行」导出的 fill_items（见 canonical_lines.fill_items_from_canonical_lines），
    以保证 Excel 列 J（报价产品规）与草稿一致。fill_items 每项需含 row, code, quote_name, unit_price, qty, specification；
    写入列 J 时使用传入的 specification（即应由调用方传入规范行中的 quote_spec 或 specification）。

    Args:
        file_path: 原始报价单路径
        fill_items: 每项含 row, code, quote_name, unit_price, qty, specification；code="无货" 表示未匹配
        sheet_name: 工作表名，默认第一个
        output_path: 输出路径，默认覆盖原文件（建议调用方传副本路径）
        freight: 运费，默认 0
        quotation_date: 报价日期，默认当天 YYYY/MM/DD
        delivery_date: 交货日期（每行同一值），默认当天 YYYY/MM/DD

    Returns:
        {"success": bool, "output_path": str, "filled_count": int, "error": str | None}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "output_path": "", "filled_count": 0, "error": "请安装 openpyxl"}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "output_path": "", "filled_count": 0, "error": f"文件不存在: {path}"}
    out_p = Path(output_path) if output_path else path
    if not out_p.is_absolute():
        out_p = Path(os.getcwd()) / out_p
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        # 先定位「Total Excluding PPN」行，避免后续 iter_rows 与写表顺序导致样式错乱
        total_row_1based = None
        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            for cell in row:
                if TOTAL_ROW_MARKER in _cell_value(cell):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        today_str = date.today().strftime("%Y/%m/%d")
        qdate_str = (quotation_date or today_str).strip() or today_str
        ddate_str = (delivery_date or today_str).strip() or today_str

        delivery_col = _find_delivery_date_column(ws)
        quotation_date_cell = _find_quotation_date_cell(ws, total_row_1based or 0) if total_row_1based else None

        # Safe document-fill mode:
        # - only write values into target cells
        # - keep existing workbook styles as-is to avoid accidental style corruption
        #   (dashed borders, unexpected fills, etc.)
        freight_value = _to_float_or_none(freight)
        if freight_value is None:
            freight_value = 0.0

        filled = 0
        total_excluding_ppn = 0.0
        for it in fill_items:
            row_num = _to_int_or_none(it.get("row"))
            if row_num is None or row_num <= 0:
                continue
            code = it.get("code")
            if code:
                _set_cell_value_merged_safe(ws, row=row_num, col=COL_PRODUCT_NO, value=str(code))
                filled += 1
            if it.get("quote_name"):
                _set_cell_value_merged_safe(ws, row=row_num, col=COL_QUOTE_NAME, value=str(it["quote_name"]))
            up = _to_float_or_none(it.get("unit_price"))
            q = _to_int_or_none(it.get("qty"))
            if up is not None:
                _set_cell_value_merged_safe(ws, row=row_num, col=COL_UNIT_PRICE, value=up)
            if q is not None:
                _set_cell_value_merged_safe(ws, row=row_num, col=COL_QTY_OUT, value=q)
            # 报价产品规格：有 specification 用 specification，否则用 quote_name，保证该列有内容（避免留空）
            spec_val = (it.get("specification") or it.get("quote_name") or "").strip()
            _set_cell_value_merged_safe(
                ws,
                row=row_num,
                col=COL_QUOTE_SPEC,
                value=spec_val if spec_val else None,
            )
            if up is not None and q is not None and code and str(code) != "无货":
                row_total = up * q
                _set_cell_value_merged_safe(ws, row=row_num, col=COL_TOTAL, value=round(row_total, 2))
                total_excluding_ppn += row_total
            elif up is not None and q is not None and (not code or str(code) == "无货"):
                _set_cell_value_merged_safe(ws, row=row_num, col=COL_TOTAL, value=0)
            if delivery_col:
                _set_cell_value_merged_safe(ws, row=row_num, col=delivery_col, value=ddate_str)

        ppn = round(total_excluding_ppn * 0.11, 2)
        total_including = round(total_excluding_ppn + ppn + freight_value, 2)
        if total_row_1based is not None:
            _set_cell_value_merged_safe(
                ws,
                row=total_row_1based,
                col=TOTALS_VALUE_COL,
                value=round(total_excluding_ppn, 2),
            )
            _set_cell_value_merged_safe(
                ws,
                row=total_row_1based + 1,
                col=TOTALS_VALUE_COL,
                value=ppn,
            )
            _set_cell_value_merged_safe(
                ws,
                row=total_row_1based + 2,
                col=TOTALS_VALUE_COL,
                value=freight_value,
            )
            _set_cell_value_merged_safe(
                ws,
                row=total_row_1based + 3,
                col=TOTALS_VALUE_COL,
                value=total_including,
            )
        if quotation_date_cell:
            qr, qc = quotation_date_cell
            _set_cell_value_merged_safe(ws, row=qr, col=qc, value=qdate_str)
        
        # 强制裁剪列：只保留 A-R (1-18列)，删除右侧所有列（包括绿色区域）
        try:
            MAX_COL = 18  # A-R
            current_max = ws.max_column
            logger.info(f"Excel 当前列数: {current_max}, 将裁剪到: {MAX_COL}")
            
            if current_max > MAX_COL:
                delete_count = current_max - MAX_COL
                ws.delete_cols(MAX_COL + 1, delete_count)
                logger.info(f"已删除 {delete_count} 列 (从第 {MAX_COL + 1} 列开始)")
                logger.info(f"裁剪后列数: {ws.max_column}")
            else:
                logger.info(f"列数 {current_max} <= {MAX_COL}, 无需裁剪")
        except Exception as e:
            logger.error(f"裁剪列失败: {e}", exc_info=True)
        
        _normalize_sheet_view(ws)
        wb.save(out_p)
        return {"success": True, "output_path": str(out_p), "filled_count": filled, "error": None}
    except Exception as e:
        return {"success": False, "output_path": "", "filled_count": 0, "error": str(e)}


def execute_quote_tool(name: str, arguments: dict[str, Any]) -> dict[str, Any]:
    """执行报价 Agent 工具，返回 {success, result, error, ...} 或 {success, items, error, rows_count}。"""
    if name == "extract_quotation_data":
        file_path = (arguments.get("file_path") or "").strip()
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        if not file_path:
            return {"success": False, "result": "", "error": "请提供 file_path", "rows_count": 0}
        return extract_quotation_data(file_path=file_path, sheet_name=sheet_name)
    if name == "fill_quotation_sheet":
        file_path = (arguments.get("file_path") or "").strip()
        fill_items = arguments.get("fill_items") or []
        output_path = (arguments.get("output_path") or "").strip() or None
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        quotation_date = (arguments.get("quotation_date") or "").strip() or None
        delivery_date = (arguments.get("delivery_date") or "").strip() or None
        if not file_path:
            return {"success": False, "result": "", "error": "请提供 file_path"}
        if not fill_items or not isinstance(fill_items, list):
            return {"success": False, "result": "", "error": "请提供 fill_items 数组"}
        out = fill_quotation(
            file_path=file_path,
            fill_items=fill_items,
            sheet_name=sheet_name,
            output_path=output_path,
            quotation_date=quotation_date,
            delivery_date=delivery_date,
        )
        if out.get("success"):
            return {"success": True, "result": json.dumps({"filled_count": out["filled_count"], "output_path": out["output_path"]}, ensure_ascii=False), "error": None}
        return {"success": False, "result": "", "error": out.get("error", "填表失败")}
    if name == "parse_excel_smart":
        fp = (arguments.get("file_path") or "").strip()
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        max_rows = arguments.get("max_rows")
        if max_rows is None:
            max_rows = 500
        try:
            max_rows = int(max_rows)
        except (TypeError, ValueError):
            max_rows = 500
        if not fp:
            return {"success": False, "result": "", "error": "请提供 file_path", "rows_read": 0}
        out = parse_excel_smart(file_path=fp, sheet_name=sheet_name, max_rows=max_rows)
        if out.get("success"):
            return {"success": True, "result": out["result"], "error": None, "rows_read": out.get("rows_read", 0)}
        return {"success": False, "result": "", "error": out.get("error", "解析失败"), "rows_read": 0}
    if name == "edit_excel":
        fp = (arguments.get("file_path") or "").strip()
        edits = arguments.get("edits") or []
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        output_path = (arguments.get("output_path") or "").strip() or None
        if not fp:
            return {"success": False, "result": "", "error": "请提供 file_path", "output_path": ""}
        out = edit_excel(file_path=fp, edits=edits, sheet_name=sheet_name, output_path=output_path)
        if out.get("success"):
            return {"success": True, "result": out["result"], "error": None, "output_path": out.get("output_path", "")}
        return {"success": False, "result": "", "error": out.get("error", "编辑失败"), "output_path": ""}
    return {"success": False, "result": "", "error": f"未知工具: {name}", "rows_count": 0}
